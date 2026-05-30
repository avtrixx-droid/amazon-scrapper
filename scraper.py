import atexit
import json
import logging
import os
import random
import re
import shutil
import signal
import smtplib
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass, field
from datetime import date, datetime, time as dt_time, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import psutil

import undetected_chromedriver as uc
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import config

__version__ = "2.0.0"

# Module-level state — reachable by atexit and signal handlers
CHROME_TEMP_DIR: Optional[str] = None
LOCK_FILE: Optional[Path] = None


# =====================================================
# Helper models
# =====================================================


@dataclass
class ASINEntry:
    asin: str
    item_name: str = ""       # from asins.txt 2nd field
    lapcare_code: str = ""    # from asins.txt 3rd field
    category: str = ""        # from # category header above this ASIN in asins.txt


@dataclass
class ScrapeResult:
    asin: str
    product_name: str
    mrp: Optional[float]
    price: Optional[float]
    discount_percent: str
    pincode: str
    city: str
    in_stock: str
    delivery_date: str
    free_delivery: str
    seller: str
    rating: str
    reviews: str
    bsr: str
    product_url: str
    scraped_at: str
    status: str  # OK / FAILED / PINCODE_FAILED
    failure_reason: str = ""


# =====================================================
# Plain-English utilities (vendor-friendly output)
# =====================================================


def print_header() -> None:
    started = datetime.now().strftime("%d %b %Y, %I:%M %p")
    print("=" * 32)
    print("Amazon.in Product Scraper")
    print(f"Started: {started}")
    print("=" * 32)


def friendly_exit(message: str, exit_code: int = 0) -> None:
    print(message)
    raise SystemExit(exit_code)


def ensure_folders(base_dir: Path) -> Dict[str, Path]:
    logs_dir = base_dir / "logs"
    output_dir = base_dir / "output"
    progress_dir = base_dir / "progress"
    logs_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    progress_dir.mkdir(parents=True, exist_ok=True)
    return {"logs": logs_dir, "output": output_dir, "progress": progress_dir}


# =====================================================
# Chrome temp dir lifecycle (ISSUE 1 fix)
# =====================================================


def cleanup_chrome_temp() -> None:
    global CHROME_TEMP_DIR
    if CHROME_TEMP_DIR and os.path.exists(CHROME_TEMP_DIR):
        try:
            shutil.rmtree(CHROME_TEMP_DIR, ignore_errors=True)
        except Exception:
            pass


def cleanup_old_chrome_dirs() -> None:
    """Delete amzscraper_chrome_* temp dirs older than 24 hours from the system temp folder."""
    try:
        tmp = Path(tempfile.gettempdir())
        cutoff = time.time() - 86400  # 24 hours
        for d in tmp.glob("amzscraper_chrome_*"):
            try:
                if d.is_dir() and d.stat().st_mtime < cutoff:
                    shutil.rmtree(d, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass


# =====================================================
# PID lock file (ISSUE 2 fix)
# =====================================================


def acquire_lock(base_dir: Path) -> None:
    global LOCK_FILE
    LOCK_FILE = base_dir / ".scraper.lock"
    if LOCK_FILE.exists():
        try:
            pid = int(LOCK_FILE.read_text().strip())
            if psutil.pid_exists(pid):
                print(f"\nThe scraper is already running (process {pid}).")
                print("Close the other terminal window first, then try again.")
                sys.exit(1)
            else:
                LOCK_FILE.unlink()
        except (ValueError, OSError):
            try:
                LOCK_FILE.unlink()
            except Exception:
                pass
    LOCK_FILE.write_text(str(os.getpid()))
    atexit.register(release_lock)


def release_lock() -> None:
    global LOCK_FILE
    try:
        if LOCK_FILE and LOCK_FILE.exists():
            LOCK_FILE.unlink()
    except OSError:
        pass


# =====================================================
# Signal handlers (ISSUE 1 + 2 fix)
# =====================================================

# These are set up inside main() so save_progress can be referenced
_signal_progress_state: Dict = {}


def _handle_exit_signal(signum, frame) -> None:
    state = _signal_progress_state
    try:
        if state.get("progress_dir") and state.get("completed_list") is not None:
            save_progress(state["progress_dir"], state.get("done_counter", 0), state["completed_list"])
    except Exception:
        pass
    cleanup_chrome_temp()
    release_lock()
    sys.exit(0)


# =====================================================
# Log cleanup
# =====================================================


def cleanup_old_logs(logs_dir: Path, keep_days: int = 30) -> None:
    """Delete log files older than keep_days days; keep at most the 30 most recent."""
    try:
        cutoff = time.time() - keep_days * 86400
        logs = sorted(logs_dir.glob("*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
        for i, log in enumerate(logs):
            if i >= 30 or log.stat().st_mtime < cutoff:
                try:
                    log.unlink()
                except Exception:
                    pass
    except Exception:
        pass


# =====================================================
# Config + input parsing
# =====================================================


def validate_config() -> Dict[str, object]:
    min_delay = getattr(config, "MIN_DELAY", 3)
    max_delay = getattr(config, "MAX_DELAY", 8)
    if not isinstance(min_delay, (int, float)) or not isinstance(max_delay, (int, float)):
        min_delay, max_delay = 3, 8

    if min_delay < 3:
        min_delay = 3
    if max_delay < min_delay:
        max_delay = min_delay + 2

    max_retries = getattr(config, "MAX_RETRIES", 2)
    if not isinstance(max_retries, int) or max_retries < 0:
        max_retries = 2

    headless = bool(getattr(config, "HEADLESS", True))

    send_email = bool(getattr(config, "SEND_EMAIL", False))
    email_from = str(getattr(config, "EMAIL_FROM", "")).strip()
    email_password = str(getattr(config, "EMAIL_PASSWORD", "")).strip()
    email_to = str(getattr(config, "EMAIL_TO", "")).strip()
    email_subject = str(getattr(config, "EMAIL_SUBJECT", "Amazon Report - {date}")).strip()

    if send_email and (not email_from or not email_password or "your-app-password" in email_password or not email_to):
        print("Email is enabled but not fully set up. Email will be skipped for this run.")
        send_email = False

    output_folder = str(getattr(config, "OUTPUT_FOLDER", "Desktop")).strip() or "Desktop"
    output_filename = str(getattr(config, "OUTPUT_FILENAME", "Amazon_Report_{date}.xlsx")).strip()

    return {
        "MIN_DELAY": float(min_delay),
        "MAX_DELAY": float(max_delay),
        "MAX_RETRIES": int(max_retries),
        "HEADLESS": headless,
        "SEND_EMAIL": send_email,
        "EMAIL_FROM": email_from,
        "EMAIL_PASSWORD": email_password,
        "EMAIL_TO": email_to,
        "EMAIL_SUBJECT": email_subject,
        "OUTPUT_FOLDER": output_folder,
        "OUTPUT_FILENAME": output_filename,
    }


def read_asins(asins_path: Path) -> List[ASINEntry]:
    """Read ASINs from asins.txt.

    Supported format (one per line):
        ASIN[,Item Name[,Lapcare Item Code]]
    Lines starting with # are category headers — applied to all ASINs below them.
    Blank lines are ignored. Invalid ASINs are skipped with a warning.
    """
    if not asins_path.exists():
        friendly_exit("Please add ASINs to asins.txt and run again")

    entries: List[ASINEntry] = []
    current_category = ""
    try:
        for raw in asins_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("#"):
                current_category = line[1:].strip()
                continue
            line = line.split("#", 1)[0].strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            asin = parts[0]
            item_name = parts[1] if len(parts) > 1 else ""
            lapcare_code = parts[2] if len(parts) > 2 else ""
            if is_valid_asin(asin):
                entries.append(ASINEntry(
                    asin=asin,
                    item_name=item_name,
                    lapcare_code=lapcare_code,
                    category=current_category,
                ))
            else:
                print(f"Skipped invalid ASIN: {asin!r}")
    except Exception:
        friendly_exit("Could not read asins.txt. Please check the file and try again.")

    if not entries:
        friendly_exit("Please add ASINs to asins.txt and run again")

    print(f"Found {len(entries)} valid ASINs")
    return entries


def is_valid_asin(value: str) -> bool:
    v = value.strip()
    return len(v) == 10 and v.startswith("B") and v.isalnum()


def read_pincodes(base_dir: Path) -> Dict[str, str]:
    pincodes_path = base_dir / "pincodes.txt"
    pincodes: Dict[str, str] = {}

    if pincodes_path.exists():
        try:
            for raw in pincodes_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if "," in line:
                    p, c = [x.strip() for x in line.split(",", 1)]
                    if p.isdigit() and len(p) == 6 and c:
                        pincodes[p] = c
        except Exception:
            print("Could not read pincodes.txt. Using pincodes from config.py instead.")

    if not pincodes:
        cfg = getattr(config, "PINCODES", {})
        if isinstance(cfg, dict):
            for p, c in cfg.items():
                p = str(p).strip()
                c = str(c).strip()
                if p.isdigit() and len(p) == 6 and c:
                    pincodes[p] = c

    if not pincodes:
        friendly_exit("No valid pincodes found. Please add pincodes in config.py or pincodes.txt and run again.")

    cities_preview = ", ".join(list(pincodes.values())[:6])
    print(f"Using {len(pincodes)} pincodes: {cities_preview}{'...' if len(pincodes) > 6 else ''}")
    return pincodes


# =====================================================
# Progress save / resume
# =====================================================


def progress_file(progress_dir: Path) -> Path:
    return progress_dir / "progress.json"


def load_progress(progress_dir: Path) -> Dict[str, object]:
    pf = progress_file(progress_dir)
    if not pf.exists():
        return {"last_completed_index": 0, "completed_combinations": [], "timestamp": ""}
    try:
        data = json.loads(pf.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"last_completed_index": 0, "completed_combinations": [], "timestamp": ""}
        data.setdefault("last_completed_index", 0)
        data.setdefault("completed_combinations", [])
        data.setdefault("timestamp", "")
        return data
    except Exception:
        return {"last_completed_index": 0, "completed_combinations": [], "timestamp": ""}


def save_progress(progress_dir: Path, last_completed_index: int, completed: List[List[str]]) -> None:
    pf = progress_file(progress_dir)
    payload = {
        "last_completed_index": int(last_completed_index),
        "completed_combinations": completed[-10000:],
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }
    try:
        pf.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except Exception:
        pass


def ask_resume(prev_done: int, total: int) -> bool:
    print("Previous session found.")
    print(f"Completed: {prev_done}/{total}. Resume? (Y/N)")
    try:
        ans = input().strip().lower()
    except Exception:
        return True
    return ans in ("y", "yes")


# =====================================================
# Results cache — incremental crash-safe storage
# =====================================================


def save_results_cache(progress_dir: Path, results_cache: Dict[str, Dict[str, "ScrapeResult"]]) -> None:
    cache_path = progress_dir / "results_cache.json"
    try:
        data: Dict = {}
        for asin, pc_dict in results_cache.items():
            data[asin] = {}
            for pc, res in pc_dict.items():
                data[asin][pc] = {
                    "asin": res.asin,
                    "product_name": res.product_name,
                    "mrp": res.mrp,
                    "price": res.price,
                    "discount_percent": res.discount_percent,
                    "pincode": res.pincode,
                    "city": res.city,
                    "in_stock": res.in_stock,
                    "delivery_date": res.delivery_date,
                    "free_delivery": res.free_delivery,
                    "seller": res.seller,
                    "rating": res.rating,
                    "reviews": res.reviews,
                    "bsr": res.bsr,
                    "product_url": res.product_url,
                    "scraped_at": res.scraped_at,
                    "status": res.status,
                    "failure_reason": res.failure_reason,
                }
        cache_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    except Exception:
        pass


def load_results_cache(progress_dir: Path) -> Dict[str, Dict[str, "ScrapeResult"]]:
    cache_path = progress_dir / "results_cache.json"
    if not cache_path.exists():
        return {}
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
        cache: Dict[str, Dict[str, ScrapeResult]] = {}
        for asin, pc_dict in data.items():
            cache[asin] = {}
            for pc, d in pc_dict.items():
                cache[asin][pc] = ScrapeResult(
                    asin=d.get("asin", asin),
                    product_name=d.get("product_name", ""),
                    mrp=d.get("mrp"),
                    price=d.get("price"),
                    discount_percent=d.get("discount_percent", "N/A"),
                    pincode=d.get("pincode", pc),
                    city=d.get("city", ""),
                    in_stock=d.get("in_stock", ""),
                    delivery_date=d.get("delivery_date", ""),
                    free_delivery=d.get("free_delivery", ""),
                    seller=d.get("seller", ""),
                    rating=d.get("rating", ""),
                    reviews=d.get("reviews", ""),
                    bsr=d.get("bsr", "N/A"),
                    product_url=d.get("product_url", ""),
                    scraped_at=d.get("scraped_at", ""),
                    status=d.get("status", "FAILED"),
                    failure_reason=d.get("failure_reason", ""),
                )
        return cache
    except Exception:
        return {}


# =====================================================
# Logging (vendor never sees traceback)
# =====================================================


def setup_logging(logs_dir: Path, timestamp: Optional[str] = None) -> logging.Logger:
    # BUG-FIX: Accept a pre-captured timestamp so Excel and log filenames share the same tag.
    # BUG-FIX: Switched format to %Y%m%d_%H%M%S so log filename includes time component.
    date_tag = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"scraper_{date_tag}.log"

    logger = logging.getLogger("amazon_scraper")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    logger.debug("Logger initialized")
    return logger


# =====================================================
# Browser setup (undetected-chromedriver)
# =====================================================


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.140 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.130 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.6045.200 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.140 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.130 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.140 Safari/537.36",
    "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.140 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5993.118 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 12_7_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.6045.200 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.130 Safari/537.36",
]


def detect_chrome_major_version(logger: logging.Logger) -> tuple[Optional[int], Optional[str]]:
    """Returns (major_version, chrome_exe_path). Either can be None on failure."""
    candidates: List[str] = []
    if sys.platform == "darwin":
        candidates.append("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
    elif sys.platform.startswith("win"):
        candidates.extend(
            [
                os.path.join(os.environ.get("PROGRAMFILES", "C:\\Program Files"), "Google", "Chrome", "Application", "chrome.exe"),
                os.path.join(os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)"), "Google", "Chrome", "Application", "chrome.exe"),
                os.path.join(os.environ.get("LOCALAPPDATA", ""), "Google", "Chrome", "Application", "chrome.exe"),
            ]
        )
    else:
        candidates.extend(["google-chrome", "chrome", "chromium", "chromium-browser"])

    # On Windows, try the registry first — more reliable than subprocess in frozen/PyInstaller envs
    if sys.platform.startswith("win"):
        try:
            import winreg
            for hive in (winreg.HKEY_CURRENT_USER, winreg.HKEY_LOCAL_MACHINE):
                for reg_path in (
                    r"Software\Google\Chrome\BLBeacon",
                    r"Software\Wow6432Node\Google\Chrome\BLBeacon",
                ):
                    try:
                        key = winreg.OpenKey(hive, reg_path)
                        version, _ = winreg.QueryValueEx(key, "version")
                        winreg.CloseKey(key)
                        m = re.search(r"(\d+)\.", str(version))
                        if m:
                            major = int(m.group(1))
                            logger.debug(f"Registry Chrome version: {version!r} -> major={major}")
                            # Find the exe path to pass to UC
                            exe_path = next((p for p in candidates if os.path.isfile(p)), None)
                            return major, exe_path
                    except OSError:
                        continue
        except ImportError:
            pass

    for exe in candidates:
        try:
            out = subprocess.check_output([exe, "--version"], stderr=subprocess.STDOUT, text=True, timeout=5).strip()
            m = re.search(r"(\d+)\.", out)
            if m:
                major = int(m.group(1))
                logger.debug(f"Detected Chrome version output: {out!r} -> major={major}")
                return major, exe
        except Exception:
            continue

    logger.debug("Could not detect Chrome major version")
    return None, None


def build_driver(headless: bool, logger: logging.Logger, base_dir: Path, worker_id: int = 0) -> Chrome:
    global CHROME_TEMP_DIR

    width = random.randint(1280, 1920)
    height = random.randint(800, 1080)
    ua = random.choice(USER_AGENTS)

    options = uc.ChromeOptions()
    options.add_argument(f"--window-size={width},{height}")
    options.add_argument("--lang=en-IN")
    options.add_argument(f"--user-agent={ua}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    # Memory limits
    options.add_argument("--memory-pressure-off")
    options.add_argument("--js-flags=--max-old-space-size=512")
    # Block images to save bandwidth and speed up page loads
    options.add_argument("--blink-settings=imagesEnabled=false")
    try:
        options.add_experimental_option("prefs", {
            "profile.managed_default_content_settings.images": 2,
            "profile.default_content_setting_values.notifications": 2,
            "profile.managed_default_content_settings.media_stream": 2,
        })
    except Exception:
        pass  # add_experimental_option may not be supported in all UC versions

    if headless:
        options.add_argument("--headless=new")

    try:
        from undetected_chromedriver.patcher import Patcher

        cache_suffix = f"_{worker_id}" if worker_id else ""
        uc_cache = (base_dir / "progress" / f"uc_cache{cache_suffix}").resolve()
        uc_cache.mkdir(parents=True, exist_ok=True)
        Patcher.data_path = str(uc_cache)
    except Exception:
        logger.debug("Could not override undetected-chromedriver cache path (non-fatal)")

    # Use a fresh temp dir per run so Chrome never accumulates state across runs (ISSUE 1)
    temp_dir = tempfile.mkdtemp(prefix="amzscraper_chrome_")
    if worker_id == 0:
        CHROME_TEMP_DIR = temp_dir
        atexit.register(cleanup_chrome_temp)

    def start_uc() -> Chrome:
        chrome_major, chrome_exe = detect_chrome_major_version(logger)
        kwargs: dict = dict(
            options=options,
            use_subprocess=True,
            user_data_dir=temp_dir,
            version_main=chrome_major,
        )
        if chrome_exe:
            kwargs["browser_executable_path"] = chrome_exe
        return uc.Chrome(**kwargs)

    try:
        driver = start_uc()
    except WebDriverException as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        msg = str(e).lower()
        logger.exception("WebDriver failed to start")
        if "chrome binary" in msg or "chrome not reachable" in msg or "cannot find chrome" in msg:
            friendly_exit(
                "Chrome browser not found.\n"
                "Please install Chrome from google.com/chrome\n"
                "Then run the script again."
            )
        if "only supports chrome version" in msg or "session not created" in msg:
            try:
                uc_cache = (base_dir / "progress" / "uc_cache").resolve()
                if uc_cache.exists():
                    for p in uc_cache.glob("*"):
                        try:
                            if p.is_file():
                                p.unlink()
                            else:
                                shutil.rmtree(p, ignore_errors=True)
                        except Exception:
                            pass
                logger.debug("Cleared uc_cache after version mismatch; retrying once")
                retry_dir = tempfile.mkdtemp(prefix="amzscraper_chrome_")
                if worker_id == 0:
                    CHROME_TEMP_DIR = retry_dir
                options_retry = options
                kwargs_retry: dict = dict(
                    options=options_retry,
                    use_subprocess=True,
                    user_data_dir=retry_dir,
                    version_main=detect_chrome_major_version(logger)[0],
                )
                driver = uc.Chrome(**kwargs_retry)
                return driver
            except Exception:
                logger.debug("Retry after cache clear failed")
            friendly_exit(
                "Chrome could not start because the ChromeDriver version does not match your Chrome version.\n"
                "Please update Google Chrome (Help → About Google Chrome), then run again.\n"
                "If Chrome was just updated, restart your computer once and try again."
            )
        if "403" in msg or "forbidden" in msg or "tunnel connection failed" in msg or "proxy" in msg:
            friendly_exit(
                "Could not start the browser because the driver download is being blocked.\n"
                "This usually happens on office Wi‑Fi / VPN / proxy networks.\n\n"
                "Please try:\n"
                "1) Run on a different Wi‑Fi (mobile hotspot works best)\n"
                "2) Turn off VPN (if any)\n"
                "3) Try again\n\n"
                "If you are on a work computer, ask IT to allow ChromeDriver downloads."
            )
        friendly_exit("Could not start the browser. Please try again.")
    except Exception:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.exception("Browser startup failed")
        friendly_exit(
            "Could not start the browser.\n"
            "Please check your internet connection and try again.\n"
            "If this is a work computer, security settings may be blocking the browser driver download."
        )

    # Page and script load timeouts (ISSUE 4)
    try:
        driver.set_page_load_timeout(30)
        driver.set_script_timeout(15)
    except Exception:
        logger.debug("Could not set page load timeout (non-fatal)")

    try:
        driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {"timezoneId": "Asia/Kolkata"})
    except Exception:
        logger.debug("Timezone override not applied (non-fatal)")

    try:
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    except Exception:
        logger.debug("webdriver flag override not applied (non-fatal)")

    return driver


# =====================================================
# Amazon interactions (pincode setting + scraping)
# =====================================================


def human_delay(settings: Dict[str, object]) -> None:
    time.sleep(random.uniform(float(settings["MIN_DELAY"]), float(settings["MAX_DELAY"])))


def detect_captcha(driver: Chrome) -> bool:
    """Cheap CAPTCHA probe — runs on every scrape, so avoid `driver.page_source`
    (full-DOM serialization). URL + title + a couple of selector probes catch
    every known Amazon.in CAPTCHA variant we've seen."""
    try:
        url = (driver.current_url or "").lower()
        if "captcha" in url or "/errors/validatecaptcha" in url:
            return True
        title = (driver.title or "").lower()
        if "captcha" in title or "robot check" in title:
            return True
        # The CAPTCHA page has a distinctive input id and a form action.
        if driver.find_elements(By.ID, "captchacharacters"):
            return True
        if driver.find_elements(By.CSS_SELECTOR, "form[action='/errors/validateCaptcha']"):
            return True
    except Exception:
        return False
    return False


def pause_for_captcha(
    driver: Chrome,
    logger: logging.Logger,
    progress_dir: Path,
    idx: int,
    completed: List[List[str]],
    msg_queue=None,
) -> None:
    """Wait up to 5 minutes for a CAPTCHA to clear, polling every 10s.

    Emits a countdown so the operator (or GUI) doesn't think the scraper hung.
    If ``msg_queue`` is provided, countdown lines are pushed there instead of
    printed — this keeps the GUI progress strip alive in worker mode.
    Returns once the CAPTCHA clears. Raises SystemExit if it persists past
    the deadline (progress is saved first).
    """
    total_seconds = 5 * 60
    poll_interval = 10
    deadline = time.time() + total_seconds

    logger.warning("CAPTCHA detected; polling every %ds for up to %ds", poll_interval, total_seconds)
    print(
        "Amazon asked for verification.\n"
        "Pausing up to 5 minutes automatically.\n"
        "Please do not close this window."
    )

    while True:
        remaining = int(deadline - time.time())
        if remaining <= 0:
            break

        mins, secs = divmod(remaining, 60)
        line = (
            f"⏳ Amazon asked for verification. Waiting up to 5 min before retry... "
            f"({mins}:{secs:02d} remaining)"
        )
        if msg_queue is not None:
            try:
                msg_queue.put_nowait(
                    {"type": "progress", "worker": -1, "done": -1, "total": -1,
                     "status": "CAPTCHA", "msg": line}
                )
            except Exception:
                pass
        else:
            print(line)

        time.sleep(min(poll_interval, max(1, remaining)))
        if not detect_captcha(driver):
            logger.info("CAPTCHA cleared after %ds", total_seconds - remaining)
            return

    # Still blocked after the full wait — save and exit cleanly.
    save_progress(progress_dir, idx, completed)
    print(
        "Script paused due to Amazon verification.\n"
        f"Progress has been saved. Will resume from ASIN {idx}."
    )
    raise SystemExit(0)


def open_homepage(driver: Chrome, logger: logging.Logger) -> None:
    for attempt in range(3):
        try:
            driver.get("https://www.amazon.in/")
            logger.debug("Opened homepage")
            return
        except Exception:
            logger.exception("Failed to open homepage")
            time.sleep(2 + attempt)
    friendly_exit("Could not open Amazon.in. Please check your internet and try again.")


def set_pincode(driver: Chrome, pincode: str, city: str, logger: logging.Logger) -> bool:
    wait = WebDriverWait(driver, 10)

    for attempt in range(1, 4):
        logger.debug(f"Setting pincode attempt {attempt}: {pincode} ({city})")
        try:
            for click_try in range(2):
                try:
                    loc_btn = wait.until(EC.element_to_be_clickable((By.ID, "nav-global-location-popover-link")))
                    loc_btn.click()
                    break
                except TimeoutException:
                    logger.debug("Location popup not clickable, retrying click")
            try:
                wait.until(EC.presence_of_element_located((By.ID, "GLUXZipUpdateInput")))
            except TimeoutException:
                logger.debug("Location popup did not appear; retrying")
                continue

            inp = driver.find_element(By.ID, "GLUXZipUpdateInput")
            inp.click()
            inp.send_keys(Keys.COMMAND if sys.platform == "darwin" else Keys.CONTROL, "a")
            inp.send_keys(Keys.BACKSPACE)
            for ch in pincode:
                inp.send_keys(ch)
                time.sleep(0.1)

            applied = False
            for sel in [
                (By.CSS_SELECTOR, "#GLUXZipUpdate .a-button-input"),
                (By.CSS_SELECTOR, "#GLUXZipUpdate input[type='submit']"),
                (By.ID, "GLUXZipUpdate"),
            ]:
                try:
                    btn = driver.find_element(*sel)
                    btn.click()
                    applied = True
                    break
                except Exception:
                    continue
            if not applied:
                logger.debug("Apply button not found; retrying")
                continue

            try:
                time.sleep(0.5)
                close_btn = driver.find_element(By.CSS_SELECTOR, "#GLUXConfirmClose")
                close_btn.click()
            except Exception:
                pass

            time.sleep(1.5)
            nav_text = ""
            try:
                nav_text = driver.find_element(By.ID, "glow-ingress-line2").text.strip()
            except Exception:
                try:
                    nav_text = driver.find_element(By.CSS_SELECTOR, "#nav-global-location-slot").text.strip()
                except Exception:
                    nav_text = ""

            logger.debug(f"Nav location text after pincode: {nav_text!r}")
            if city.lower() in nav_text.lower() or pincode in nav_text:
                return True

            try:
                err = driver.find_element(By.CSS_SELECTOR, "#GLUXZipError .a-alert-content").text.strip()
                if err:
                    logger.debug(f"Pincode error: {err}")
            except Exception:
                pass

        except Exception:
            logger.exception("Pincode set failed")

        time.sleep(2.0)

    logger.warning(f"Pincode not accepted after 3 tries: {pincode}")
    return False


def asin_url(asin: str) -> str:
    return f"https://www.amazon.in/dp/{asin}"


def safe_get_text(driver: Chrome, selectors: List[Tuple[By, str]], logger: logging.Logger) -> str:
    for by, sel in selectors:
        try:
            el = driver.find_element(by, sel)
            # .text returns empty for CSS-hidden elements (a-offscreen uses position:absolute left:-9999px)
            # get_attribute("textContent") always returns the raw text regardless of visibility
            txt = (el.text or el.get_attribute("textContent") or "").strip()
            logger.debug(f"Selector success: {by} {sel} -> {txt[:120]!r}")
            if txt:
                return txt
        except NoSuchElementException:
            logger.debug(f"Selector not found: {by} {sel}")
        except Exception:
            logger.exception(f"Selector error: {by} {sel}")
    return "Not Found"


def parse_money(text: str) -> Optional[float]:
    if not text or text == "Not Found":
        return None
    cleaned = text.replace("₹", "").replace(",", "").strip()
    m = re.search(r"(\d+(?:\.\d+)?)", cleaned)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def extract_price(driver: Chrome, logger: logging.Logger) -> Optional[float]:
    """Extract current selling price. Tries specific containers first to avoid
    picking up sponsored or related-product prices from generic .a-price-whole elements."""

    # Strategy 1: specific price containers (modern Amazon India structure)
    # Use textContent on .a-offscreen because .text returns empty for hidden elements
    specific_offscreen = [
        "#corePriceDisplay_desktop_feature_div .apexPriceToPay span.a-offscreen",
        "#corePriceDisplay_desktop_feature_div .priceToPay span.a-offscreen",
        "#corePrice_feature_div .apexPriceToPay span.a-offscreen",
        "#corePrice_feature_div .priceToPay span.a-offscreen",
        ".apexPriceToPay span.a-offscreen",
        ".priceToPay span.a-offscreen",
    ]
    for sel in specific_offscreen:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            txt = (el.get_attribute("textContent") or "").strip()
            val = parse_money(txt)
            if val and val > 0:
                logger.debug(f"Price via offscreen {sel}: {val}")
                return val
        except NoSuchElementException:
            continue
        except Exception:
            logger.debug(f"Price selector error: {sel}")

    # Strategy 2: whole + fraction from specific containers
    whole_containers = [
        "#corePriceDisplay_desktop_feature_div",
        "#corePrice_feature_div",
        ".apexPriceToPay",
        ".priceToPay",
        "",  # bare fallback — any .a-price-whole on page
    ]
    for container in whole_containers:
        whole_sel = f"{container} .a-price-whole".strip()
        try:
            whole_el = driver.find_element(By.CSS_SELECTOR, whole_sel)
            whole = (whole_el.text or whole_el.get_attribute("textContent") or "").strip()
            if not whole:
                continue
            frac = "00"
            try:
                frac_sel = f"{container} .a-price-fraction".strip()
                frac_el = driver.find_element(By.CSS_SELECTOR, frac_sel)
                frac = (frac_el.text or frac_el.get_attribute("textContent") or "00").strip()
            except Exception:
                pass
            combined = f"{whole}.{frac}".replace(",", "").replace(" ", "").strip(".")
            val = parse_money(combined)
            if val and val > 0:
                logger.debug(f"Price via whole+frac ({container or 'bare'}): {val}")
                return val
        except NoSuchElementException:
            continue
        except Exception:
            logger.debug(f"Price whole+frac error for container: {container!r}")

    # Strategy 3: legacy selectors (older Amazon pages)
    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, "#priceblock_dealprice"),
            (By.CSS_SELECTOR, "#priceblock_ourprice"),
            (By.CSS_SELECTOR, "#price_inside_buybox"),
            (By.CSS_SELECTOR, "#tp_price_block_total_price_ww span.a-offscreen"),
        ],
        logger,
    )
    return parse_money(txt)


def extract_mrp(driver: Chrome, logger: logging.Logger) -> Optional[float]:
    """Extract MRP (original crossed-out price). Uses textContent on hidden elements."""
    # Specific container selectors first
    specific = [
        "#corePriceDisplay_desktop_feature_div .basisPrice span.a-offscreen",
        "#corePrice_feature_div .basisPrice span.a-offscreen",
        "#corePriceDisplay_desktop_feature_div .a-price.a-text-price span.a-offscreen",
        "#corePrice_feature_div .a-price.a-text-price span.a-offscreen",
    ]
    for sel in specific:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            txt = (el.get_attribute("textContent") or "").strip()
            val = parse_money(txt)
            if val and val > 0:
                logger.debug(f"MRP via {sel}: {val}")
                return val
        except NoSuchElementException:
            continue
        except Exception:
            logger.debug(f"MRP selector error: {sel}")

    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, ".basisPrice span.a-offscreen"),
            (By.CSS_SELECTOR, ".a-price.a-text-price span.a-offscreen"),
            (By.CSS_SELECTOR, "span.a-text-price > span.a-offscreen"),
            (By.CSS_SELECTOR, "#priceblock_ourprice"),
        ],
        logger,
    )
    return parse_money(txt)


def extract_availability(driver: Chrome, logger: logging.Logger) -> str:
    """
    Cross-check two Amazon availability signals.
    Returns: "In Stock" | "Out of Stock" | "Low Stock (X left)" | "Unknown"

    BUG-FIX: Previous code read only #availability span, which can say "In Stock"
    even when no seller has an active offer. Now cross-checks with buy box state.
    """

    # BUG-FIX: Signal A — Catalog availability text (multi-selector waterfall)
    AVAILABILITY_SELECTORS = [
        "#availability span",
        "#availabilityInsideBuyBox_feature_div span",
        "#outOfStock span",
        "#almAvailability_feature_div span.primary-availability-message",  # BUG-FIX: Amazon Now
    ]
    availability_text = ""  # BUG-FIX
    for sel in AVAILABILITY_SELECTORS:  # BUG-FIX
        try:  # BUG-FIX
            el = driver.find_element(By.CSS_SELECTOR, sel)  # BUG-FIX
            text = (el.text or el.get_attribute("textContent") or "").strip().lower()  # BUG-FIX
            if text:  # BUG-FIX
                availability_text = text  # BUG-FIX
                break  # BUG-FIX
        except Exception:  # BUG-FIX
            continue  # BUG-FIX

    # BUG-FIX: Signal B — Can the product actually be purchased right now?
    BUY_BOX_ACTIVE_SELECTORS = [
        "#add-to-cart-button",
        "#buy-now-button",
        "#freshAddToCartButton",          # BUG-FIX: Amazon Now Add to Cart
        "input[name='submit.add-to-cart']",
        "#submit.add-to-cart-ubb",
    ]
    buy_box_active = False  # BUG-FIX
    for sel in BUY_BOX_ACTIVE_SELECTORS:  # BUG-FIX
        try:  # BUG-FIX
            el = driver.find_element(By.CSS_SELECTOR, sel)  # BUG-FIX
            if el.is_displayed():  # BUG-FIX
                buy_box_active = True  # BUG-FIX
                break  # BUG-FIX
        except Exception:  # BUG-FIX
            continue  # BUG-FIX

    # BUG-FIX: Signal C — Explicit "currently unavailable" phrases on page
    UNAVAILABLE_PHRASES = [
        "currently unavailable",
        "not available",
        "we don't know when or if",
        "sign up to be notified",
        "item under review",
        "this item cannot be shipped",  # BUG-FIX: common for restricted products
    ]
    page_says_unavailable = False  # BUG-FIX
    try:  # BUG-FIX
        page_text = driver.find_element(By.ID, "availability").text.lower()  # BUG-FIX
        page_says_unavailable = any(p in page_text for p in UNAVAILABLE_PHRASES)  # BUG-FIX
    except Exception:  # BUG-FIX
        pass  # BUG-FIX: Element missing is OK — handled by Signal B

    logger.debug(  # BUG-FIX
        f"Availability signals — text={availability_text!r} buy_box={buy_box_active} "  # BUG-FIX
        f"unavailable_phrase={page_says_unavailable}"  # BUG-FIX
    )  # BUG-FIX

    # BUG-FIX: Rule 1 — Page explicitly says unavailable → Out of Stock
    if page_says_unavailable:  # BUG-FIX
        return "Out of Stock"  # BUG-FIX

    # BUG-FIX: Rule 2 — No buy box button visible → Out of Stock (key fix)
    if not buy_box_active:  # BUG-FIX
        return "Out of Stock"  # BUG-FIX

    # BUG-FIX: Rule 3 — Low stock indicator
    if "only" in availability_text and "left" in availability_text:  # BUG-FIX
        m = re.search(r'only (\d+) left', availability_text)  # BUG-FIX
        if m:  # BUG-FIX
            return f"Low Stock ({m.group(1)} left)"  # BUG-FIX

    # BUG-FIX: Rule 4 — Both signals agree it's available
    if "in stock" in availability_text and buy_box_active:  # BUG-FIX
        return "In Stock"  # BUG-FIX

    # BUG-FIX: Rule 5 — Ambiguous text but buy box present
    if buy_box_active:  # BUG-FIX
        return "In Stock"  # BUG-FIX

    return "Out of Stock"  # BUG-FIX


_DELIVERY_NOISE_RE = re.compile(
    r"^(or\s+)?fastest\s+delivery\s+"
    # Don't strip the "in" / "by" that follows "delivery" — the duration and
    # date parsers need it ("FREE delivery in 10 minutes" -> "in 10 minutes").
    r"|^free\s+delivery\s+"
    r"|^get\s+it\s+(by\s+)?"
    r"|^delivery\s+by\s+"
    r"|\.\s*order\s+within[\s\d\w]+"
    r"|on\s+your\s+first\s+order\.?"
    r"|limited\s+period\s+offer\.?"
    r"|\[?details\]?",
    re.IGNORECASE,
)

_MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
# BUG-FIX-AMZNOW: include long-form month names so "June 2" / "Delivery by
# Tuesday, June 4" parse correctly. Previously only 3-letter prefixes were in
# the alternation and `\b` after "jun" failed against "june", so explicit dates
# with full month names silently fell through to weekday-name inference and
# produced the wrong delivery date.
_MONTH_PAT = (
    r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may"
    r"|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?"
    r"|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
)
_WEEKDAYS = [
    "monday", "tuesday", "wednesday", "thursday",
    "friday", "saturday", "sunday",
]


def _strip_delivery_noise(text: str) -> str:
    """Strip Amazon's prefixes/suffixes so the bare date phrase is left."""
    cleaned = _DELIVERY_NOISE_RE.sub("", text)
    return re.sub(r"\s{2,}", " ", cleaned).strip(" .,")


def _resolve_month_day(now: datetime, month: int, day: int) -> Optional[datetime]:
    """Place a (month, day) onto a concrete year — current, or next when more
    than ~60 days in the past so a Jan date seen in Dec rolls forward correctly."""
    try:
        candidate = datetime(now.year, month, day, 12, 0)
    except ValueError:
        return None
    if (candidate - now).days < -60:
        try:
            candidate = datetime(now.year + 1, month, day, 12, 0)
        except ValueError:
            return None
    return candidate


def _parse_delivery_phrase(raw_text: str, now: Optional[datetime] = None) -> Optional[datetime]:
    """Convert a delivery phrase into a concrete delivery datetime.

    Returns None if no date can be extracted. Earlier datetimes mean earlier
    deliveries. The caller sorts by this directly — no minute-bucketing.

    Strategy in order of confidence:
      1. Durations  ("in 10 minutes" / "10 Minutes" / "2 hours")
      2. Day spans  ("in 3 days" / "2-3 days" / "3 Days")
      3. Explicit dates ("20 May" / "May 20") — most precise, tried before
         "tomorrow" so "Tomorrow, 20 May" lands on the actual 20th.
      4. Relative  ("today" / "tomorrow")
      5. Weekday names ("Thursday")

    Note: durations and day-counts match WITH OR WITHOUT a leading "in" because
    `_pick_earliest_result` re-parses the display string ("Amazon Now – 10
    Minutes (Free)") where `_build_delivery_display` already stripped the
    "free delivery in " prefix. Requiring "in" here would silently drop
    Amazon Now options from the global earliest pick.
    """
    if not raw_text:
        return None
    if now is None:
        now = datetime.now()
    today = now.date()

    t = _strip_delivery_noise(raw_text.lower())

    # 1. Durations — bare or "in N <unit>"
    # BUG-FIX-AMZNOW: abbreviations "min"/"mins"/"hr"/"hrs" were silently
    # dropped (parser only matched the full word "minute"/"hour"). Amazon Now
    # widgets frequently render "in 30 mins" / "in 2 hrs", so this hid the
    # fastest channel — the candidate was discarded and the slower Standard
    # row won. Use a single word-boundary alternation that covers all forms.
    m = re.search(r"\b(\d+)\s+(?:minutes?|mins?)\b", t)
    if m:
        return now + timedelta(minutes=int(m.group(1)))
    m = re.search(r"\b(\d+)\s+(?:hours?|hrs?)\b", t)
    if m:
        return now + timedelta(hours=int(m.group(1)))

    # 2. Day spans — range first ("2-3 days" → low bound), then bare/in-prefixed
    m = re.search(r"(\d+)\s*-\s*\d+\s+days?\b", t)
    if m:
        return datetime.combine(today + timedelta(days=int(m.group(1))), dt_time(12, 0))
    m = re.search(r"\b(\d+)\s+days?\b", t)
    if m:
        return datetime.combine(today + timedelta(days=int(m.group(1))), dt_time(12, 0))

    # 2b. Date ranges with month ("8-9 May") — pick the LOW end. Without this,
    # Step 3's single-date regex eats only the second number and we'd silently
    # report the slower end of the window.
    m = re.search(
        rf"\b(\d{{1,2}})\s*-\s*\d{{1,2}}\s+{_MONTH_PAT}\b",
        t,
    )
    if m:
        dt = _resolve_month_day(now, _MONTHS[m.group(2)], int(m.group(1)))
        if dt:
            return dt

    # 3. Explicit date — try BEFORE "tomorrow"/weekday so a phrase like
    #    "Tomorrow, 20 May" lands on May 20 even if today isn't quite May 19.
    m = re.search(rf"\b(\d{{1,2}})\s+{_MONTH_PAT}\b", t)
    if m:
        dt = _resolve_month_day(now, _MONTHS[m.group(2)], int(m.group(1)))
        if dt:
            return dt
    m = re.search(rf"\b{_MONTH_PAT}\s+(\d{{1,2}})\b", t)
    if m:
        dt = _resolve_month_day(now, _MONTHS[m.group(1)], int(m.group(2)))
        if dt:
            return dt

    # 4. Relative day
    # BUG-FIX-AMZNOW: "tomorrow" → tomorrow at noon (well before any specific
    # courier promise on that day). "today" → today at 20:00 (end of business)
    # rather than noon, so a "10 minutes" / "2 hours" candidate captured at,
    # say, 14:00 still sorts AHEAD of the generic "today" promise. Previously
    # "today" at noon was already in the past for any afternoon scrape, which
    # caused it to win against same-day minutes/hours candidates.
    if re.search(r"\btomorrow\b", t):
        return datetime.combine(today + timedelta(days=1), dt_time(12, 0))
    if re.search(r"\btoday\b", t):
        return datetime.combine(today, dt_time(20, 0))

    # 5. Weekday name
    today_idx = today.weekday()
    for idx, name in enumerate(_WEEKDAYS):
        if re.search(rf"\b{name}\b", t):
            ahead = (idx - today_idx) % 7
            if ahead == 0:
                ahead = 7
            return datetime.combine(today + timedelta(days=ahead), dt_time(12, 0))

    return None


def _normalise_delivery_to_minutes(channel: str, raw_text: str) -> int:
    """Backward-compat shim: minutes-from-now for legacy callers/log lines.
    The real sort is on the datetime returned by _parse_delivery_phrase."""
    now = datetime.now()
    dt = _parse_delivery_phrase(raw_text, now)
    if dt is None:
        return 999999
    delta = int((dt - now).total_seconds() // 60)
    return max(0, delta)


def _clean_delivery_text(raw_text: str) -> str:
    """Strip Amazon's prefixes/suffixes/noise; return the bare promise text.

    Examples:
        "FREE delivery in 10 minutes on orders over ₹149" → "10 Minutes"
        "Or fastest delivery Tomorrow, 31 May. Details"   → "Tomorrow, 31 May"
        "Get it by Monday, June 3"                        → "Monday, June 3"
    """
    text = (raw_text or "").strip()
    # Leading noise — "Or fastest delivery …", "FREE delivery in …", "Get it by …"
    text = re.sub(r"^(or\s+)?fastest\s+delivery\s+", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(
        r"^(free\s+delivery\s+in\s+|free\s+delivery\s+|get\s+it\s+(by\s+)?|delivery\s+by\s+|delivery\s+in\s+|delivery\s+)",
        "", text, flags=re.IGNORECASE,
    ).strip()
    # Trailing noise — "on orders over ₹149", "Details", countdown, badges
    text = re.sub(r"\s+on orders over.*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\.\s*order within[\s\d\w]+", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"on\s+your\s+first\s+order\.?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"limited\s+period\s+offer\.?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\[?details\]?\.?\s*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\s{2,}", " ", text).strip(" .,")
    return text.title()


def _build_delivery_display(channel: str, raw_text: str, is_free: bool) -> str:
    """Legacy display builder: 'Amazon Now – 10 Minutes (Free)'.

    Kept for backward compatibility with `extract_earliest_delivery_from_text`
    and the test suite. The live extractor uses `_build_amazon_now_display`
    and `_build_standard_display` directly (no (Free) suffix, no channel
    prefix for Standard rows — the user-confirmed format).
    """
    text = _clean_delivery_text(raw_text)
    free_label = " (Free)" if is_free else ""
    return f"{channel} – {text}{free_label}"


def _build_amazon_now_display(raw_text: str) -> str:
    """Tier 1 output: 'Amazon Now – 10 Minutes' / 'Amazon Now – 2 Hours'."""
    text = _clean_delivery_text(raw_text)
    return f"Amazon Now – {text}" if text else ""


def _build_standard_display(raw_text: str) -> str:
    """Tier 2 output: 'Tomorrow, 31 May' / 'Today' / '2 June' (no prefix, no (Free))."""
    return _clean_delivery_text(raw_text)


# ── Pattern-based selectors for the tiered delivery extractor ────────────────
# SENIOR-FIX-TIERED: Instead of hardcoding every exact slot ID Amazon ships
# (`…_PRIMARY_DELIVERY_MESSAGE_LARGE`, `…_SECONDARY_…_SMALL`, etc.), we match
# the *prefix patterns* with CSS substring selectors. The day Amazon ships a
# `…_TERTIARY_…` slot, we pick it up automatically — no code change. The `i`
# flag makes the match case-insensitive (Amazon mixes `DELIVERY_BLOCK` upper
# and `deliveryBlockMessage` camel in different layouts).

# Tier 1 — Amazon Now: ALM, quick-commerce buy-box row, ALM offer display.
_AMAZON_NOW_SELECTOR = (
    "[id*='alm-delivery' i],"
    "[id*='qcomBuyBox' i],"
    "[id*='almOfferDisplay' i]"
)

# Tier 2 — Standard delivery: MIR slots (PRIMARY/SECONDARY/PROMISE/…),
# legacy single-message containers, promise holder.
_STANDARD_DELIVERY_SELECTOR = (
    "[id*='DELIVERY_BLOCK' i],"        # MIR layout — any slot suffix
    "[id*='deliveryBlock' i],"          # legacy #deliveryBlockMessage
    "[id*='deliveryMessage' i],"        # #ddmDeliveryMessage / #deliveryMessageMirWidget
    "[id*='promiseMessage' i]"          # #promiseMessage_feature_div
)

# For HTML-dump forensics when all tiers fail.
_DEBUG_DUMP_CONTAINERS = [
    "alm-delivery-message",
    "qcomBuyBoxRow_feature_div",
    "almOfferDisplay_feature_div",
    "mir-layout-DELIVERY_BLOCK",
    "deliveryBlockMessage",
    "ddmDeliveryMessage",
    "deliveryMessageMirWidget",
    "promiseMessage_feature_div",
    "freshDeliveryMessage_feature_div",
    "mbc",
    "newAccordionRow_0",
    "newAccordionRow_1",
    "buybox",
    "rightCol",
]

# Phrase patterns scanned out of free-text. Any match is fed to
# `_parse_delivery_phrase`; non-parseable matches are dropped silently.
_DATE_PHRASE_RE = re.compile(
    r"(?:FREE\s+)?(?:Or\s+fastest\s+)?delivery\s+(?:in\s+)?"
    r"(?:\d+\s+(?:minutes?|mins?|hours?|hrs?|days?)"
    r"|today\b[^.<\"\n]{0,80}"
    r"|tomorrow\b[^.<\"\n]{0,80}"
    r"|(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b[^.<\"\n]{0,80}"
    r"|\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^.<\"\n]{0,30})",
    re.IGNORECASE,
)

# Fallback patterns when there's no leading "delivery" keyword — picks up bare
# "Today, 12 PM" / "Tomorrow, 20 May" / "Thursday, 22 May" inside slot text.
_BARE_DATE_PHRASE_RE = re.compile(
    r"\b(?:today|tomorrow|monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b"
    r"[^.<\"\n]{0,80}",
    re.IGNORECASE,
)

# Duration patterns ("10 minutes", "2 hours") that don't need the word "delivery"
# preceding them — common in Amazon Now widgets.
_BARE_DURATION_RE = re.compile(
    r"\b\d+\s+(?:minutes?|mins?|hours?|hrs?)\b[^.<\"\n]{0,60}",
    re.IGNORECASE,
)


def _delivery_debug_dir(logger: logging.Logger) -> Optional[Path]:
    """Derive a forensics dump folder from the logger's file handler.

    Returns <log-dir>/delivery_debug/ or None if the logger has no file handler.
    """
    try:
        for h in logger.handlers:
            base = getattr(h, "baseFilename", None)
            if base:
                d = Path(base).resolve().parent / "delivery_debug"
                d.mkdir(parents=True, exist_ok=True)
                return d
    except Exception:
        pass
    return None


def _dump_delivery_html(
    driver: Chrome, asin_hint: str, pincode: str, dump_dir: Path, reason: str
) -> None:
    """Save a snapshot of every known delivery container for forensic post-mortem.

    Called when zero candidates are found OR at INFO/DEBUG cadence on first hit
    per pincode. File is text — opens in any editor.
    """
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_asin = re.sub(r"[^A-Za-z0-9_-]", "_", asin_hint or "noasin")
        out = dump_dir / f"{safe_asin}_{pincode}_{ts}_{reason}.html"
        chunks = [f"<!-- reason={reason} pincode={pincode} ts={ts} -->\n"]
        for cid in _DEBUG_DUMP_CONTAINERS:
            try:
                el = driver.find_element(By.ID, cid)
                html = el.get_attribute("outerHTML") or ""
                chunks.append(f"\n<!-- ===== #{cid} ===== -->\n{html}\n")
            except Exception:
                continue
        # Also dump the contextual pincode label so we can verify what Amazon
        # thinks the active pincode is at the moment of extraction.
        try:
            label = driver.find_element(By.ID, "contextualIngressPtLabel_deliveryShortLine")
            chunks.append(
                f"\n<!-- ===== #contextualIngressPtLabel_deliveryShortLine ===== -->\n"
                f"{label.get_attribute('outerHTML') or ''}\n"
            )
        except Exception:
            pass
        out.write_text("".join(chunks), encoding="utf-8")
    except Exception:
        pass  # Never let dumping crash a scrape


def _infer_channel(text: str, hint: Optional[str]) -> str:
    """Best-guess channel name from a delivery fragment."""
    if hint:
        return hint
    low = text.lower()
    if "amazon fresh" in low or "amazonfresh" in low:
        return "Amazon Fresh"
    if "amazon now" in low or re.search(r"\b\d+\s*(?:minutes?|mins?|hours?|hrs?)\b", low):
        return "Amazon Now"
    if re.search(r"\bprime\b", low):
        return "Prime"
    return "Standard"


def _is_free(text: str) -> bool:
    low = text.lower()
    return bool(re.search(r"\bfree\b|₹\s*0\b|no\s+charge", low))


# BUG-FIX-AMZNOW: Pure-Python earliest-delivery resolver.
# Selenium-free so it's unit-testable; called from `extract_all_delivery_options`
# for the page-source / container-text path AND used directly by the test suite
# in tests/test_delivery_parser.py.
def extract_earliest_delivery_from_text(
    text: str,
    default_channel: Optional[str] = None,
    now: Optional[datetime] = None,
) -> dict:
    """Scan a blob of text for every delivery promise and return the earliest.

    Used by both the live scraper (to consolidate the page-source / container
    fallback) and the unit tests (to validate priority ordering without spinning
    up Chrome).

    Priority is enforced *implicitly* by the datetime each candidate parses to:
      minutes-based  →  now + N min       (earliest)
      hours-based    →  now + N hours
      same-day/today →  today at noon
      tomorrow       →  tomorrow at noon
      explicit date  →  that date at noon
      weekday name   →  next occurrence at noon
      unparseable    →  dropped

    Returns ``{"earliest_display", "earliest_dt", "is_free", "all_options"}``.
    On no parseable match, ``earliest_display`` is "Not Available".
    """
    if now is None:
        now = datetime.now()
    result: dict = {
        "earliest_display": "Not Available",
        "earliest_dt": None,
        "is_free": False,
        "all_options": [],
    }
    if not text:
        return result

    candidates: List[Tuple[str, str, bool, datetime]] = []
    seen_phrases: set = set()

    def _line_around(idx: int) -> str:
        # BUG-FIX-AMZNOW: channel inference used to run only on the regex match
        # itself, so "Amazon Fresh Delivery in 2 hours" inferred Standard because
        # the captured "Delivery in 2 hours" had no channel keyword. Look at the
        # enclosing line for hints like "Amazon Fresh" / "Prime".
        start = text.rfind("\n", 0, idx) + 1
        end = text.find("\n", idx)
        if end == -1:
            end = len(text)
        return text[start:end]

    def _ingest(match_start: int, raw: str) -> None:
        raw = re.sub(r"\s+", " ", raw or "").strip(" .,")
        if not raw or len(raw) < 4:
            return
        low = raw.lower()
        # Skip fragments that obviously aren't delivery promises
        if not any(kw in low for kw in (
            "today", "tomorrow", "min", "hour", "day", "week",
            "jan", "feb", "mar", "apr", "may", "jun",
            "jul", "aug", "sep", "oct", "nov", "dec",
            "monday", "tuesday", "wednesday", "thursday",
            "friday", "saturday", "sunday",
        )):
            return
        dt = _parse_delivery_phrase(raw, now)
        if dt is None:
            return
        key = (raw.lower(), dt)
        if key in seen_phrases:
            return
        seen_phrases.add(key)
        # Channel inference: try (default → line context → fragment text).
        ch = default_channel or _infer_channel(_line_around(match_start), None)
        if ch == "Standard":
            # Last attempt — maybe the raw fragment itself has the channel keyword.
            ch = _infer_channel(raw, None)
        candidates.append((ch, raw, _is_free(raw), dt))

    # Pass 1 — phrases anchored on the word "delivery"
    for m in _DATE_PHRASE_RE.finditer(text):
        _ingest(m.start(), m.group(0))
    # Pass 2 — bare durations ("10 minutes", "2 hrs") which Amazon Now widgets
    # render without a leading "delivery" keyword
    for m in _BARE_DURATION_RE.finditer(text):
        _ingest(m.start(), m.group(0))
    # Pass 3 — bare date phrases ("Tomorrow, 31 May", "Monday, June 3")
    for m in _BARE_DATE_PHRASE_RE.finditer(text):
        _ingest(m.start(), m.group(0))

    if not candidates:
        return result

    candidates.sort(key=lambda c: c[3])
    # Dedup by (channel, lowercased display string) so "FREE delivery in 10
    # minutes on orders over ₹149" and the bare "10 minutes on orders over ₹149"
    # don't both appear in the option list.
    deduped: List[Tuple[str, str, bool, datetime]] = []
    seen_display: set = set()
    for ch, raw, free, dt in candidates:
        disp_key = (ch, _build_delivery_display(ch, raw, free).lower())
        if disp_key in seen_display:
            continue
        seen_display.add(disp_key)
        deduped.append((ch, raw, free, dt))

    earliest_ch, earliest_raw, earliest_free, earliest_dt = deduped[0]
    result["earliest_display"] = _build_delivery_display(
        earliest_ch, earliest_raw, earliest_free
    )
    result["earliest_dt"] = earliest_dt
    result["is_free"] = earliest_free
    result["all_options"] = [
        {
            "channel": ch,
            "raw_text": raw,
            "display_text": _build_delivery_display(ch, raw, free),
            "delivery_dt": dt,
            "is_free": free,
        }
        for ch, raw, free, dt in deduped
    ]
    return result


# ── Buy-box anchoring ────────────────────────────────────────────────────────
# SENIOR-FIX-BUYBOX: Delivery extraction MUST be scoped to the buy box (the
# right-hand offer panel that holds Price / Availability / Quantity / Add to
# Cart / Buy Now / Delivery / Sold by / Ships from / Payment). Reading delivery
# text from the whole document picked up the word "today" / "tomorrow" / weekday
# names from customer reviews ("today my package arrived"), the product
# description ("Get yours today!"), Q&A ("Will this arrive Monday?"), and
# "Frequently bought together" widgets — every one a false positive that beat
# the real promise on `datetime` sort order.
#
# Anchor strategy:
#   1. Try strict buy-box wrapper IDs in order of specificity
#   2. Validate each candidate by *scoring* the markers it contains
#      (CTA buttons, price element, quantity, sold-by tabular, delivery block)
#   3. Pick the highest-scoring candidate; require score >= 2
#   4. Fall back to #rightCol (broader but always a strict superset of the buy
#      box on desktop product pages)

# Candidates from most-specific to broadest. `#buybox` is the actual offer
# panel; `#rightCol` always contains it on desktop product pages.
_BUY_BOX_CANDIDATE_IDS: list = [
    "buybox",
    "desktop_buybox",
    "apex_desktop",
    "rightCol",
    "centerCol_feature_div",
    "centerCol",
]

# Buttons that exist *only* in the buy box. Strongest signal we have.
_BUY_BOX_CTA_SELECTOR = (
    "#add-to-cart-button, #buy-now-button, "
    "input[name='submit.add-to-cart'], input[name='submit.buy-now'], "
    "#one-click-button, #buyNow_feature_div input"
)

# Price element selectors — these can appear in recommendation widgets too,
# so on their own they aren't proof of the buy box.
_BUY_BOX_PRICE_SELECTOR = (
    "#corePriceDisplay_desktop_feature_div, #corePrice_feature_div, "
    "#priceblock_ourprice, #priceblock_dealprice, #priceblock_saleprice, "
    "#price_inside_buybox"
)

# Buy-box-only structural markers.
_BUY_BOX_AUX_SELECTORS = [
    "#quantity, #quantitySelect, [data-action='a-dropdown-button']",
    "#tabular-buybox, .tabular-buybox-text, #merchant-info, #sellerProfileTriggerId",
    "[id*='DELIVERY_BLOCK'], #alm-delivery-message, "
    "#freshDeliveryMessage_feature_div, #qcomBuyBoxRow_feature_div, "
    "#almOfferDisplay_feature_div, #mbc, [id^='newAccordionRow_']",
    "#availability, #outOfStock, #exports-desktop-out-of-stock-message",
]


def _score_buy_box_candidate(el) -> int:
    """Score 0–6 by counting buy-box markers present inside an element.

    Score interpretation:
      0–1 → almost certainly not the buy box (or empty wrapper)
      2–3 → likely buy box
      4+ → definitely buy box
    """
    try:
        score = 0
        if el.find_elements(By.CSS_SELECTOR, _BUY_BOX_CTA_SELECTOR):
            score += 2  # CTA buttons are the strongest single signal
        if el.find_elements(By.CSS_SELECTOR, _BUY_BOX_PRICE_SELECTOR):
            score += 1
        for sel in _BUY_BOX_AUX_SELECTORS:
            if el.find_elements(By.CSS_SELECTOR, sel):
                score += 1
        return score
    except Exception:
        return 0


def find_buy_box(driver: Chrome, logger: logging.Logger):
    """Locate the buy box. Returns the WebElement or ``None`` if not found.

    SENIOR-FIX-BUYBOX: every delivery extraction call must use this anchor —
    otherwise we read review/Q&A/description text and produce wrong dates.
    """
    best_el = None
    best_score = 0
    best_id = None
    for cid in _BUY_BOX_CANDIDATE_IDS:
        try:
            el = driver.find_element(By.ID, cid)
        except Exception:
            continue
        score = _score_buy_box_candidate(el)
        if score > best_score:
            best_el = el
            best_score = score
            best_id = cid

    if best_score >= 2:
        logger.debug(
            f"Buy box anchored on #{best_id} (score={best_score}/6)"
        )
        return best_el

    if best_el is not None:
        logger.warning(
            f"Buy box best candidate #{best_id} only scored {best_score}/6 — "
            f"delivery extraction may include off-target text."
        )
        return best_el

    logger.warning(
        "No buy box candidate found on page — delivery extraction will be "
        "unscoped and may pick up reviews/description text."
    )
    return None


# BUG-FIX-DEALBADGE: Amazon's buy box renders promotional badges that share
# keywords with delivery phrases — "Today's Deal", "Limited Time Deal",
# "Deal of the Day", etc. The bare-date regex matches "today" inside
# "Today's Deal" and the parser resolves it to today's date — producing a
# false positive like "Standard – Today'S Deal (Free)" in the Excel.
#
# Defense: in Tier 3 (buy-box innerText sweep), only accept lines that
# contain a delivery-context keyword. Tier 1 and Tier 2 are already scoped
# to delivery-specific containers (`alm-delivery`, `DELIVERY_BLOCK`, etc.)
# so they can't see badge text.
_DELIVERY_CONTEXT_RE = re.compile(
    r"\bdelivery\b"
    r"|\bdeliver\b"
    r"|\barriv(?:e|es|ing)\b"
    r"|\bship(?:s|ping|ped)?\b"
    r"|\bget\s+it\b"
    r"|\bfastest\b"
    r"|\bfree\s+(?:delivery|shipping)\b"
    r"|\bby\s+(?:today|tomorrow|mon|tue|wed|thu|fri|sat|sun"
    r"|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b",
    re.IGNORECASE,
)


def _filter_delivery_lines(text: str) -> str:
    """Keep only the lines of `text` that contain a delivery-context keyword.

    Used by Tier 3 to skip promotional badges like "Today's Deal" / "Limited
    Time Deal" that share words with our date regex.
    """
    if not text:
        return ""
    keep: list = []
    for line in text.splitlines():
        if _DELIVERY_CONTEXT_RE.search(line):
            keep.append(line)
    return "\n".join(keep)


def _element_text(el) -> str:
    """Read an element's visible text, normalising whitespace. Returns ''."""
    try:
        raw = el.text or el.get_attribute("textContent") or ""
    except Exception:
        return ""
    return re.sub(r"\s+", " ", raw).strip()


def _extract_amazon_now(
    scope, now: datetime, logger: logging.Logger, asin_hint: str, pincode: str
) -> Optional[dict]:
    """Tier 1: Amazon Now / Quick Commerce. Trust Amazon — if any quick-commerce
    container has parseable text inside the buy box, return it as the earliest.

    Matches IDs containing `alm-delivery`, `qcomBuyBox`, or `almOfferDisplay`
    (case-insensitive). Also walks up from any `img[alt*='Amazon Now']` to its
    nearest container so we catch image-anchored variants.
    """
    elements: list = []
    try:
        elements.extend(scope.find_elements(By.CSS_SELECTOR, _AMAZON_NOW_SELECTOR))
    except Exception as e:
        logger.debug(f"[{asin_hint}][{pincode}] Tier 1 ID scan failed: {e}")

    try:
        for img in scope.find_elements(By.CSS_SELECTOR, "img[alt*='Amazon Now' i]"):
            try:
                container = img.find_element(By.XPATH, "./ancestor::div[1]")
                elements.append(container)
            except Exception:
                continue
    except Exception as e:
        logger.debug(f"[{asin_hint}][{pincode}] Tier 1 image walk failed: {e}")

    for el in elements:
        text = _element_text(el)
        if not text:
            continue
        dt = _parse_delivery_phrase(text, now)
        if dt is None:
            continue
        display = _build_amazon_now_display(text)
        if not display:
            continue
        is_free_flag = _is_free(text)
        logger.info(
            f"[{asin_hint}][{pincode}] EARLIEST (Tier 1 / Amazon Now): "
            f"{display!r} @ {dt} | raw={text!r}"
        )
        return {
            "earliest_display": display,
            "is_free": is_free_flag,
            "all_options": [{
                "channel": "Amazon Now",
                "raw_text": text,
                "display_text": display,
                "delivery_dt": dt,
                "is_free": is_free_flag,
            }],
        }
    return None


def _extract_standard_delivery(
    scope, now: datetime, logger: logging.Logger, asin_hint: str, pincode: str
) -> Optional[dict]:
    """Tier 2: Standard delivery — MIR slots (PRIMARY/SECONDARY/PROMISE/…) and
    legacy single-message containers. Read all matches, parse every phrase
    inside each, return the earliest by datetime.

    PRIMARY is usually FREE+slower; SECONDARY is usually paid+faster
    ("Or fastest delivery …"). Picking earliest gives the vendor the fastest
    available option; the price flag is surfaced separately in the Free
    Delivery column.
    """
    try:
        elements = scope.find_elements(By.CSS_SELECTOR, _STANDARD_DELIVERY_SELECTOR)
    except Exception as e:
        logger.debug(f"[{asin_hint}][{pincode}] Tier 2 selector scan failed: {e}")
        return None

    candidates: list = []
    seen_texts: set = set()
    for el in elements:
        text = _element_text(el)
        if not text or text in seen_texts:
            continue
        seen_texts.add(text)
        # Each element may contain multiple promises (e.g. #deliveryBlockMessage
        # wraps both PRIMARY and SECONDARY children) — enumerate every phrase.
        text_result = extract_earliest_delivery_from_text(text, "Standard", now)
        for opt in text_result["all_options"]:
            candidates.append(opt)

    if not candidates:
        return None

    candidates.sort(key=lambda c: c["delivery_dt"])
    # Dedup by formatted display so PRIMARY's "Tomorrow, 31 May" and the same
    # phrase repeated in the parent block don't both occupy a slot.
    deduped: list = []
    seen_display: set = set()
    for c in candidates:
        display = _build_standard_display(c["raw_text"])
        key = display.lower()
        if key in seen_display:
            continue
        seen_display.add(key)
        deduped.append({
            "channel": "Standard",
            "raw_text": c["raw_text"],
            "display_text": display,
            "delivery_dt": c["delivery_dt"],
            "is_free": c["is_free"],
        })

    earliest = deduped[0]
    logger.info(
        f"[{asin_hint}][{pincode}] EARLIEST (Tier 2 / Standard): "
        f"{earliest['display_text']!r} @ {earliest['delivery_dt']} "
        f"| candidates={len(deduped)}"
    )
    return {
        "earliest_display": earliest["display_text"],
        "is_free": earliest["is_free"],
        "all_options": deduped,
    }


def _extract_buybox_fallback(
    buy_box, now: datetime, logger: logging.Logger, asin_hint: str, pincode: str
) -> Optional[dict]:
    """Tier 3: Last-resort regex sweep on the buy box's innerText.

    Only runs when Tiers 1+2 returned nothing — Amazon must have shipped a new
    container ID we don't recognise. Logged as a WARNING so the failure is
    visible and the new selector can be added to Tier 1/2 in a follow-up.
    """
    try:
        inner = buy_box.get_attribute("innerText") or buy_box.text or ""
    except Exception:
        return None
    if not inner:
        return None

    # BUG-FIX-DEALBADGE: drop lines without a delivery keyword so the bare-date
    # regex doesn't match "Today's Deal" / "Limited Time Deal" / "Deal of the
    # Day" promotional badges that sit next to the price in the buy box.
    filtered = _filter_delivery_lines(inner)
    if not filtered:
        logger.warning(
            f"[{asin_hint}][{pincode}] Tier 3: buy-box innerText has no "
            f"delivery-context lines — skipping to avoid badge false positives."
        )
        return None

    text_result = extract_earliest_delivery_from_text(filtered, None, now)
    if text_result["earliest_dt"] is None:
        return None

    earliest_opt = text_result["all_options"][0]
    if earliest_opt["channel"] == "Amazon Now":
        display = _build_amazon_now_display(earliest_opt["raw_text"])
    else:
        display = _build_standard_display(earliest_opt["raw_text"])

    logger.warning(
        f"[{asin_hint}][{pincode}] EARLIEST (Tier 3 / buy-box fallback): "
        f"{display!r} @ {earliest_opt['delivery_dt']} — "
        f"Tiers 1+2 found nothing; consider adding a new selector."
    )
    return {
        "earliest_display": display,
        "is_free": earliest_opt["is_free"],
        "all_options": [{
            "channel": earliest_opt["channel"],
            "raw_text": earliest_opt["raw_text"],
            "display_text": display,
            "delivery_dt": earliest_opt["delivery_dt"],
            "is_free": earliest_opt["is_free"],
        }],
    }


def extract_all_delivery_options(
    driver: Chrome,
    expected_pincode: str,
    logger: logging.Logger,
    asin_hint: str = "",
) -> dict:
    """Tiered, buy-box-scoped delivery extractor.

    The strategy mirrors Amazon's own DOM hierarchy — each tier maps to a
    distinct fulfilment family. First tier that returns a parseable date wins.

      A. Verify pincode applied                       (global — header label)
      B. Wait for delivery widget to populate         (global)
      C. Locate the buy box                           (anchor; everything below is scoped to it)
      D. Tier 1 — Amazon Now / Quick Commerce
            [id*='alm-delivery' i], [id*='qcomBuyBox' i], [id*='almOfferDisplay' i]
            + img[alt*='Amazon Now']. Trust Amazon — if rendered, it's the fastest.
      E. Tier 2 — Standard delivery
            [id*='DELIVERY_BLOCK' i], [id*='deliveryBlock' i],
            [id*='deliveryMessage' i], [id*='promiseMessage' i].
            Read all matches, parse every phrase, return earliest by datetime.
      F. Tier 3 — Buy-box innerText regex sweep (insurance)
            Runs only if Tiers 1+2 returned nothing. Logged as WARNING so the
            new container ID can be added to Tier 1/2 in a follow-up.
      G. Tier 4 — "Not Available"  (dump HTML for forensics)

    Output format (user-confirmed):
      • Tier 1 hit  → "Amazon Now – 10 Minutes" / "Amazon Now – 2 Hours"
      • Tier 2/3 hit → "Tomorrow, 31 May" / "Today" / "2 June"  (no channel
                         prefix, no "(Free)" suffix — price is in a separate
                         Excel column)
      • Tier 4      → "Not Available"

    Returns:
        {
            "earliest_display": str,
            "is_free":          bool,
            "all_options":      list,
            "pincode_verified": bool,
            "buy_box_found":    bool,
            "tier":             "amazon_now" | "standard" | "fallback" | "none",
        }
    """
    result: dict = {
        "earliest_display": "Not Available",
        "is_free": False,
        "all_options": [],
        "pincode_verified": False,
        "buy_box_found": False,
        "tier": "none",
    }
    debug_dir = _delivery_debug_dir(logger)
    now = datetime.now()

    # ── STEP A. Verify pincode actually applied ───────────────────────────────
    try:
        WebDriverWait(driver, 12).until(
            lambda d: expected_pincode in (
                d.find_element(By.ID, "contextualIngressPtLabel_deliveryShortLine").text or ""
            )
        )
        result["pincode_verified"] = True
        logger.info(f"[{asin_hint}][{expected_pincode}] Pincode confirmed on product page.")
    except Exception:
        logger.warning(
            f"[{asin_hint}][{expected_pincode}] Could not verify pincode label — "
            f"delivery data may be stale. Proceeding with 3s grace sleep."
        )
        time.sleep(3)

    # ── STEP B. Wait for delivery widget to render with non-empty content ────
    # PERF-FIX: dropped from 15s → 8s. The wait condition is "any delivery
    # widget populated" — 8s is generous; pages that haven't rendered by then
    # have rendered nothing, and we fall through to Tier 4 ("Not Available").
    delivery_ready = wait_for_delivery_block(driver, timeout=8)
    if not delivery_ready:
        logger.warning(
            f"[{asin_hint}][{expected_pincode}] Delivery widget did not populate "
            f"within 8s — dumping HTML for post-mortem."
        )
        if debug_dir:
            _dump_delivery_html(driver, asin_hint, expected_pincode, debug_dir, "widget_empty")

    # ── STEP C. Anchor on the buy box ────────────────────────────────────────
    buy_box = find_buy_box(driver, logger)
    result["buy_box_found"] = buy_box is not None
    # When we can't find a buy box, fall back to driver-wide selector reads.
    # The [id*='…' i] patterns are still safe outside the buy box (they only
    # match delivery-prefixed IDs), but we skip the Tier 3 innerText sweep
    # without an anchor — that's the one that risks review/Q&A false positives.
    scope = buy_box if buy_box is not None else driver

    # ── STEP D. Tier 1 — Amazon Now / Quick Commerce ─────────────────────────
    tier1 = _extract_amazon_now(scope, now, logger, asin_hint, expected_pincode)
    if tier1 is not None:
        result.update(tier1)
        result["tier"] = "amazon_now"
        return result

    # ── STEP E. Tier 2 — Standard delivery (MIR slots + legacy) ──────────────
    tier2 = _extract_standard_delivery(scope, now, logger, asin_hint, expected_pincode)
    if tier2 is not None:
        result.update(tier2)
        result["tier"] = "standard"
        return result

    # ── STEP F. Tier 3 — Last-resort buy-box innerText sweep ─────────────────
    if buy_box is not None:
        tier3 = _extract_buybox_fallback(buy_box, now, logger, asin_hint, expected_pincode)
        if tier3 is not None:
            result.update(tier3)
            result["tier"] = "fallback"
            return result

    # ── STEP G. Tier 4 — Not Available (dump HTML for forensics) ─────────────
    logger.warning(
        f"[{asin_hint}][{expected_pincode}] No delivery promise found by any "
        f"tier. Dumping delivery container HTML for forensics."
    )
    if debug_dir:
        _dump_delivery_html(driver, asin_hint, expected_pincode, debug_dir, "zero_candidates")
    return result


def extract_product_name(driver: Chrome, logger: logging.Logger) -> str:
    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, "#productTitle"),
            (By.CSS_SELECTOR, ".product-title"),
            (By.CSS_SELECTOR, "h1.a-size-large"),
        ],
        logger,
    )
    return re.sub(r"\s+", " ", txt).strip()


def extract_seller(driver: Chrome, logger: logging.Logger) -> str:
    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, "#merchant-info a"),
            (By.CSS_SELECTOR, "#sellerProfileTriggerId"),
            (By.CSS_SELECTOR, ".offer-display-feature-text"),
        ],
        logger,
    )
    if txt == "Not Found":
        return "Amazon"
    return re.sub(r"\s+", " ", txt).strip()


def extract_rating(driver: Chrome, logger: logging.Logger) -> str:
    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, "#acrPopover span.a-icon-alt"),
            (By.CSS_SELECTOR, '[data-hook="rating-out-of-text"]'),
        ],
        logger,
    )
    m = re.search(r"(\d+(?:\.\d+)?)", txt)
    return m.group(1) if m else "Not Found"


def extract_review_count(driver: Chrome, logger: logging.Logger) -> str:
    txt = safe_get_text(
        driver,
        [
            (By.CSS_SELECTOR, "#acrCustomerReviewText"),
            (By.CSS_SELECTOR, '[data-hook="total-review-count"]'),
        ],
        logger,
    )
    m = re.search(r"([\d,]+)", txt)
    if not m:
        return "Not Found"
    return m.group(1).replace(",", "")


def extract_bsr(driver: Chrome, logger: logging.Logger) -> str:
    """Extract Best Sellers Rank (Product Ranking) from Amazon product page."""
    bsr_selectors = [
        (By.CSS_SELECTOR, "#SalesRank"),
        (By.CSS_SELECTOR, "#productDetails_detailBullets_sections1"),
        (By.CSS_SELECTOR, "#detailBulletsWrapper_feature_div"),
        (By.CSS_SELECTOR, "#productDetails_db_sections"),
        (By.CSS_SELECTOR, ".a-section.a-spacing-small.a-padding-small"),
    ]
    for by, sel in bsr_selectors:
        try:
            el = driver.find_element(by, sel)
            text = el.text or ""
            lower = text.lower()
            if "best seller" in lower or "best-seller" in lower or "#" in text:
                m = re.search(r"#\s*([\d,]+)\s+in\s+([^\n(]{3,60})", text)
                if m:
                    rank = m.group(1).replace(",", "")
                    category = re.sub(r"\s+", " ", m.group(2)).strip().rstrip(".")
                    logger.debug(f"BSR found via {sel}: #{rank} in {category}")
                    return f"#{rank} in {category}"
                m2 = re.search(r"#\s*([\d,]+)", text)
                if m2:
                    rank = m2.group(1).replace(",", "")
                    logger.debug(f"BSR rank only via {sel}: #{rank}")
                    return f"#{rank}"
        except NoSuchElementException:
            continue
        except Exception:
            logger.debug(f"BSR extraction error for {sel}")
            continue

    # Last resort: page source scan
    try:
        page = driver.page_source or ""
        m = re.search(r"#\s*([\d,]+)\s+in\s+([^<\n]{3,60}?)(?:<|\n|&)", page)
        if m:
            rank = m.group(1).replace(",", "")
            category = re.sub(r"\s+", " ", m.group(2)).strip().rstrip(".")
            logger.debug(f"BSR from page source: #{rank} in {category}")
            return f"#{rank} in {category}"
    except Exception:
        pass

    return "N/A"


def compute_discount(mrp: Optional[float], price: Optional[float]) -> str:
    if not mrp or not price or mrp <= 0:
        return "N/A"
    try:
        return str(int(round((mrp - price) / mrp * 100)))
    except Exception:
        return "N/A"


def random_backoff(seconds_min: int, seconds_max: int) -> None:
    time.sleep(random.uniform(seconds_min, seconds_max))


def wait_for_product_page(driver: Chrome, timeout: int = 20) -> bool:
    """Wait until #productTitle is visible — avoids fixed sleep after navigation."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.ID, "productTitle"))
        )
        return True
    except Exception:
        return False


# BUG-FIX: Wait for page to confirm new pincode before reading delivery date.
# Without this wait, the previous pincode's delivery date is still showing.
def wait_for_pincode_confirmation(driver: Chrome, expected_pincode: str, timeout: int = 12) -> bool:
    """
    Block until Amazon's page confirms the new pincode is active.
    Returns True if confirmed within timeout, False otherwise.

    BUG-FIX: This is the core fix for delivery dates copying across pincodes.
    The delivery block refreshes asynchronously — we must wait for it.
    """
    PINCODE_CONFIRMATION_SELECTORS = [  # BUG-FIX
        "contextualIngressPtLabel_deliveryShortLine",  # BUG-FIX: primary (from live HTML)
        "glow-ingress-line2",                          # BUG-FIX: fallback — header location text
        "nav-global-location-data-modal-action",       # BUG-FIX: fallback — nav bar location
    ]

    for element_id in PINCODE_CONFIRMATION_SELECTORS:  # BUG-FIX
        try:  # BUG-FIX
            WebDriverWait(driver, timeout).until(  # BUG-FIX
                lambda d, eid=element_id, pc=expected_pincode: (  # BUG-FIX
                    pc in d.find_element(By.ID, eid).text  # BUG-FIX
                )  # BUG-FIX
            )  # BUG-FIX
            logging.getLogger("amazon_scraper").debug(  # BUG-FIX
                f"Pincode {expected_pincode} confirmed via #{element_id}"  # BUG-FIX
            )  # BUG-FIX
            return True  # BUG-FIX
        except Exception:  # BUG-FIX
            continue  # BUG-FIX: try next selector

    logging.getLogger("amazon_scraper").warning(  # BUG-FIX
        f"BUG-FIX WARNING: Could not confirm pincode {expected_pincode} on page. "  # BUG-FIX
        f"Delivery date for this pincode may be inaccurate. "  # BUG-FIX
        f"Falling back to 5s sleep."  # BUG-FIX
    )  # BUG-FIX
    time.sleep(5)  # BUG-FIX
    return False  # BUG-FIX


def wait_for_delivery_block(driver: Chrome, timeout: int = 15) -> bool:
    """Wait until ANY delivery element has rendered with non-empty text.

    Critical: Amazon's delivery widget is Ajax-driven. The slot element appears
    in the DOM almost immediately but stays empty for ~1–3s after a pincode
    change while the new courier promise is fetched. Waiting only on element
    presence (the old behavior) read a blank/stale widget and silently produced
    "Not Available" or the previous pincode's date.

    This wait polls for ANY of these conditions, all at once (not sequentially):
      - `[data-csa-c-delivery-time]` element exists  (DEX unified widget)
      - any `[id*='DELIVERY_BLOCK-slot']` has non-empty text  (MIR slots — covers
        PRIMARY, SECONDARY, PROMISE_HOLDER, UB, and any future variant)
      - `#alm-delivery-message` has non-empty text  (Amazon Now / ALM)
      - `#deliveryBlockMessage` or `#ddmDeliveryMessage` has non-empty text
        (legacy widgets still seen on some product categories)
    """
    # PERF-FIX: Reverted the overstrict "wait for fast-channel containers when
    # present" condition. Amazon ships #mbc / #newAccordionRow_* / #qcomBuyBox*
    # as EMPTY PLACEHOLDERS on most product pages even when there's no Amazon
    # Now / Fresh offer for that pincode. Waiting for them to populate timed
    # out at the full 15s on every scrape, costing ~90s per ASIN on 6 pincodes.
    # We now wait only for "any one delivery widget populated" — same as the
    # original. The tiered extractor handles the case where Amazon Now wasn't
    # populated yet (Tier 1 returns None, Tier 2 takes over).
    def _ready(d) -> bool:
        try:
            js = """
              return (function(){
                var sels = [
                  '[data-csa-c-delivery-time]',
                  '[id*=\"DELIVERY_BLOCK-slot\"]',
                  '#alm-delivery-message',
                  '#deliveryBlockMessage',
                  '#ddmDeliveryMessage',
                  '#deliveryMessageMirWidget'
                ];
                for (var i = 0; i < sels.length; i++) {
                  var nodes = document.querySelectorAll(sels[i]);
                  for (var j = 0; j < nodes.length; j++) {
                    var t = (nodes[j].innerText || nodes[j].textContent || '').trim();
                    if (t.length > 0) return true;
                    if (sels[i] === '[data-csa-c-delivery-time]') {
                      var a = nodes[j].getAttribute('data-csa-c-delivery-time') || '';
                      if (a.trim().length > 0) return true;
                    }
                  }
                }
                return false;
              })();
            """
            return bool(d.execute_script(js))
        except Exception:
            return False
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.4).until(_ready)
        return True
    except Exception:
        return False


def validate_page_is_product(driver: Chrome, asin: str) -> str:
    """Return 'OK', 'CAPTCHA', 'NOT_FOUND', 'WRONG_PAGE', or 'INCOMPLETE_LOAD'.

    Uses URL + title + selector probes instead of ``driver.page_source`` —
    avoids a full-DOM serialization on every scrape (3000-row run = 3000
    serializations otherwise).
    """
    try:
        url = (driver.current_url or "").lower()
        title = (driver.title or "").lower()
    except Exception:
        return "WRONG_PAGE"

    if "captcha" in url or "/errors/validatecaptcha" in url or "captcha" in title or "robot check" in title:
        return "CAPTCHA"
    if "page-not-found" in url or "/404" in url or "page not found" in title:
        return "NOT_FOUND"

    # Canonical product URL contains the ASIN — if neither URL nor title has
    # any Amazon footprint, we clearly aren't where we expected.
    if asin.lower() not in url and "amazon" not in title:
        return "WRONG_PAGE"

    # Page rendered if any of these key elements exist. They're all O(1) lookups.
    try:
        if driver.find_elements(By.ID, "productTitle"):
            return "OK"
        if driver.find_elements(By.ID, "add-to-cart-button"):
            return "OK"
        if driver.find_elements(By.ID, "outOfStock"):
            return "OK"
    except Exception:
        return "WRONG_PAGE"
    return "INCOMPLETE_LOAD"


_SCRAPE_COUNT = 0
_RECYCLE_EVERY = 50  # Recycle Chrome every N successful scrapes to bound memory.


def check_browser_health(driver: Chrome, logger: logging.Logger) -> Tuple[bool, bool]:
    """Inspect the driver and decide whether the caller should recycle it.

    Returns ``(is_alive, should_recycle)``:
      - ``is_alive``  – False means the driver has crashed; caller MUST restart.
      - ``should_recycle`` – True every ``_RECYCLE_EVERY`` calls as a pre-emptive
        hint to swap drivers before Chrome memory creeps. The caller controls
        the actual restart so it can re-set pincode / re-warm afterwards.
    """
    global _SCRAPE_COUNT
    _SCRAPE_COUNT += 1
    is_alive = True
    try:
        _ = driver.current_url
    except Exception:
        logger.warning("Browser unresponsive — will recycle")
        is_alive = False
    should_recycle = (_SCRAPE_COUNT % _RECYCLE_EVERY == 0)
    return is_alive, should_recycle


def _recycle_driver(
    driver: Optional[Chrome],
    settings: Dict[str, object],
    logger: logging.Logger,
    base_dir: Path,
    worker_id: int = 0,
) -> Chrome:
    """Quit the current driver, free its temp dir, and return a fresh driver.

    Safe to call with a None / dead driver — quit failures are swallowed.
    Caller is responsible for re-doing anything session-scoped (pincode, etc.).
    """
    try:
        if driver is not None:
            driver.quit()
    except Exception:
        logger.debug("Old driver quit raised (non-fatal)")
    cleanup_chrome_temp()
    new_driver = build_driver(
        bool(settings["HEADLESS"]), logger, base_dir, worker_id=worker_id
    )
    open_homepage(new_driver, logger)
    return new_driver


def scrape_one(driver: Chrome, asin: str, pincode: str, city: str, logger: logging.Logger) -> ScrapeResult:
    url = asin_url(asin)
    scraped_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    logger.debug(f"Visiting URL: {url}")
    driver.get(url)

    # Wait for product title to appear (adaptive — no fixed sleep)
    wait_for_product_page(driver, timeout=20)
    # PERF-FIX: removed redundant wait_for_delivery_block() here — it's already
    # called inside extract_all_delivery_options(). Calling it twice was costing
    # up to 15s per scrape (×6 pincodes ≈ 90s wasted per ASIN).

    # Validate the page before extracting
    page_state = validate_page_is_product(driver, asin)
    if page_state == "CAPTCHA" or detect_captcha(driver):
        return ScrapeResult(
            asin=asin,
            product_name="",
            mrp=None,
            price=None,
            discount_percent="",
            pincode=pincode,
            city=city,
            in_stock="",
            delivery_date="",
            free_delivery="",
            seller="",
            rating="",
            reviews="",
            bsr="",
            product_url=url,
            scraped_at=scraped_at,
            status="FAILED",
            failure_reason="CAPTCHA",
        )
    if page_state == "NOT_FOUND":
        return ScrapeResult(
            asin=asin, product_name="Not Found", mrp=None, price=None,
            discount_percent="N/A", pincode=pincode, city=city,
            in_stock="Not Found", delivery_date="Not Available", free_delivery="N/A",
            seller="", rating="", reviews="", bsr="N/A",
            product_url=url, scraped_at=scraped_at, status="FAILED",
            failure_reason="Product not found (404)",
        )

    product_name = extract_product_name(driver, logger)
    mrp = extract_mrp(driver, logger)
    price = extract_price(driver, logger)
    discount = compute_discount(mrp, price)
    availability = extract_availability(driver, logger)
    delivery_result = extract_all_delivery_options(driver, pincode, logger, asin_hint=asin)
    delivery = delivery_result["earliest_display"]
    free_delivery = "Yes" if delivery_result["is_free"] else "No"
    if not delivery_result["pincode_verified"]:
        logger.warning(f"[{asin}][{pincode}] Delivery read without pincode confirmation — data may be stale.")
    seller = extract_seller(driver, logger)
    rating = extract_rating(driver, logger)
    reviews = extract_review_count(driver, logger)
    bsr = extract_bsr(driver, logger)

    return ScrapeResult(
        asin=asin,
        product_name=product_name,
        mrp=mrp,
        price=price,
        discount_percent=discount,
        pincode=pincode,
        city=city,
        in_stock=availability,
        delivery_date=delivery,
        free_delivery=free_delivery,
        seller=seller,
        rating=rating,
        reviews=reviews,
        bsr=bsr,
        product_url=url,
        scraped_at=scraped_at,
        status="OK",
    )


def scrape_with_smart_retry(
    driver: Chrome,
    asin: str,
    pincode: str,
    city: str,
    logger: logging.Logger,
    progress_dir: Path,
    idx_display: int,
    total_combos: int,
    completed_list: List[List[str]],
    settings: Dict[str, object],
    base_dir: Path,
) -> Tuple[ScrapeResult, Chrome]:
    """Run scrape_one with up to 3 attempts. After any CAPTCHA pause the driver
    is recycled (CLAUDE.md: 'do NOT reuse a stale driver instance'). Returns
    ``(result, driver)`` so the caller picks up the possibly-fresh driver."""

    def _after_captcha(d: Chrome) -> Chrome:
        d = _recycle_driver(d, settings, logger, base_dir)
        if set_pincode(d, pincode, city, logger):
            wait_for_pincode_confirmation(d, pincode)
        return d

    res = scrape_one(driver, asin, pincode, city, logger)
    if res.failure_reason == "CAPTCHA":
        pause_for_captcha(driver, logger, progress_dir, idx_display, completed_list)
        driver = _after_captcha(driver)
        res = scrape_one(driver, asin, pincode, city, logger)

    if res.status == "OK":
        print(
            f"✅ [{idx_display:04d}/{total_combos}] {asin} | {city:<10} | "
            f"₹{int(res.price) if res.price else 'NA'} | {res.delivery_date[:12]:<12} | {res.in_stock}"
        )
        return res, driver

    print(f"❌ [{idx_display:04d}/{total_combos}] {asin} | {city:<10} | Failed - Retrying...")

    # Attempt 2
    random_backoff(5, 15)
    try:
        driver.refresh()
    except Exception:
        pass
    res2 = scrape_one(driver, asin, pincode, city, logger)
    if res2.failure_reason == "CAPTCHA":
        pause_for_captcha(driver, logger, progress_dir, idx_display, completed_list)
        driver = _after_captcha(driver)
        res2 = scrape_one(driver, asin, pincode, city, logger)
    if res2.status == "OK":
        print(
            f"✅ [{idx_display:04d}/{total_combos}] {asin} | {city:<10} | "
            f"₹{int(res2.price) if res2.price else 'NA'} | {res2.delivery_date[:12]:<12} | {res2.in_stock}"
        )
        return res2, driver

    print(f"❌ [{idx_display:04d}/{total_combos}] {asin} | {city:<10} | Failed - Retrying...")

    # Attempt 3 (final)
    random_backoff(30, 35)
    try:
        driver.refresh()
    except Exception:
        pass
    res3 = scrape_one(driver, asin, pincode, city, logger)
    if res3.failure_reason == "CAPTCHA":
        pause_for_captcha(driver, logger, progress_dir, idx_display, completed_list)
        driver = _after_captcha(driver)
        res3 = scrape_one(driver, asin, pincode, city, logger)
    if res3.status == "OK":
        print(
            f"✅ [{idx_display:04d}/{total_combos}] {asin} | {city:<10} | "
            f"₹{int(res3.price) if res3.price else 'NA'} | {res3.delivery_date[:12]:<12} | {res3.in_stock}"
        )
        return res3, driver

    res3.status = "FAILED"
    res3.failure_reason = "Failed after retries"
    return res3, driver


# =====================================================
# Excel output — pivoted (one row per ASIN)
# =====================================================


# Fixed columns before the dynamic per-pincode columns.
# Order matches CLAUDE.md "Excel Output Schema" section (cols A–N + Scraped At).
FIXED_HEADERS = [
    "Category",            # A
    "Item Name",           # B
    "Lapcare Item Code",   # C
    "ASIN Link",           # D
    "Current Price (₹)",   # E
    "MRP (₹)",             # F
    "Discount (%)",        # G
    "Product Ranking",     # H
    "Availability",        # I — overall availability across pincodes
    "Rating",              # J
    "Reviews",             # K
    "Seller",              # L
    "Earliest Delivery",   # M — fastest delivery across all pincodes
    "Free Delivery",       # N — Yes/No on the earliest option
    "Scraped At",          # before the per-pincode block
]

# Column index (1-based) of the ASIN Link in FIXED_HEADERS
_ASIN_LINK_COL = 4


def get_desktop_path() -> Path:
    home = Path.home()
    return home / "Desktop"


def resolve_output_path(settings: Dict[str, object], timestamp: Optional[str] = None) -> Path:
    # BUG-FIX: Accept a pre-captured timestamp so Excel and log filenames share the same tag.
    # BUG-FIX: Switched format to %Y%m%d_%H%M%S per spec (e.g. 20240514_143022).
    date_tag = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = str(settings["OUTPUT_FILENAME"]).replace("{date}", date_tag)

    folder_raw = str(settings["OUTPUT_FOLDER"]).strip()
    if folder_raw.lower() == "desktop":
        folder = get_desktop_path()
    else:
        folder = Path(folder_raw).expanduser()
        if not folder.is_absolute():
            folder = (Path.cwd() / folder).resolve()

    try:
        folder.mkdir(parents=True, exist_ok=True)
        probe = folder / ".write_test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)  # type: ignore[arg-type]
        return folder / filename
    except Exception:
        # When frozen by PyInstaller, __file__ points inside _MEIPASS temp dir;
        # write user data next to the executable instead.
        if getattr(sys, "frozen", False):
            fallback = Path(sys.executable).resolve().parent / "output"
        else:
            fallback = Path(__file__).resolve().parent / "output"
        fallback.mkdir(parents=True, exist_ok=True)
        print(f"Could not write to '{folder}'. Saving report to: {fallback}")
        return fallback / filename


def format_results_header(ws, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"


def format_failed_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E1"


def _pincode_cell_value(res: ScrapeResult) -> str:
    """Format the per-pincode availability cell content."""
    if res.status == "PINCODE_FAILED":
        return "Pincode Failed"
    if res.status == "FAILED":
        return "Scrape Failed"
    avail = res.in_stock or "Check on Amazon"
    delivery = res.delivery_date or ""
    if delivery and delivery not in ("Not Found", "Set pincode manually"):
        return f"{avail} | {delivery}"
    return avail


def _pick_earliest_result(
    asin_results: Dict[str, ScrapeResult],
    now: Optional[datetime] = None,
) -> Optional[ScrapeResult]:
    """Return the OK result whose delivery_date parses to the earliest datetime.

    Used to populate the "Earliest Delivery" and "Free Delivery" columns (the
    fastest option across ALL pincodes for an ASIN). Returns None if no result
    has a parseable delivery date.
    """
    if now is None:
        now = datetime.now()
    best: Optional[ScrapeResult] = None
    best_dt: Optional[datetime] = None
    for r in asin_results.values():
        if r.status != "OK":
            continue
        dt = _parse_delivery_phrase(r.delivery_date or "", now)
        if dt is None:
            continue
        if best_dt is None or dt < best_dt:
            best = r
            best_dt = dt
    return best


def _style_pivoted_row(
    ws,
    row_num: int,
    asin_results: Dict[str, ScrapeResult],
    fixed_col_count: int,
    pincodes: Dict[str, str],
) -> None:
    """Apply alternate shading and per-pincode color coding to a pivoted row."""
    total_cols = fixed_col_count + len(pincodes)
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    if row_num % 2 == 0:
        for col in range(1, total_cols + 1):
            ws.cell(row=row_num, column=col).fill = alt_fill

    # Color each pincode cell individually
    for col_idx, (pincode, _city) in enumerate(pincodes.items(), start=fixed_col_count + 1):
        res = asin_results.get(pincode)
        if res is None:
            continue
        cell = ws.cell(row=row_num, column=col_idx)
        if res.status in ("FAILED", "PINCODE_FAILED"):
            cell.fill = PatternFill("solid", fgColor="F0F0F0")
        elif (res.in_stock or "").lower().startswith("out of stock"):
            cell.fill = PatternFill("solid", fgColor="FFE0E0")
        elif (res.in_stock or "").lower().startswith(("in stock", "low stock")):
            days = estimate_delivery_days(res.delivery_date)
            if days is None:
                pass
            elif days <= 1:
                cell.fill = PatternFill("solid", fgColor="E0FFE0")
            elif days <= 3:
                cell.fill = PatternFill("solid", fgColor="FFFDE0")
            else:
                cell.fill = PatternFill("solid", fgColor="FFF0E0")


def build_pivoted_excel(
    results_cache: Dict[str, Dict[str, ScrapeResult]],
    asin_entries: List[ASINEntry],
    pincodes: Dict[str, str],
    xlsx_path: Path,
    failed_rows: List[List],
    logger: logging.Logger,
) -> Workbook:
    """Build a fresh pivoted workbook — one row per ASIN, one column per pincode."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    pincode_headers = [f"{pc} - {city}" for pc, city in pincodes.items()]
    all_headers = FIXED_HEADERS + pincode_headers
    ws.append(all_headers)
    format_results_header(ws, len(all_headers))

    for entry in asin_entries:
        asin = entry.asin
        asin_results = results_cache.get(asin, {})

        # Skip ASINs that haven't been scraped at all yet
        if not asin_results:
            continue

        # Pick the first successful result for shared fields (price, name, etc.)
        any_ok = next((r for r in asin_results.values() if r.status == "OK"), None)
        any_result = any_ok or next(iter(asin_results.values()))

        # Fastest-delivery result across all pincodes (CLAUDE.md cols M, N).
        # Falls back to `any_ok` so we still show *something* if no delivery
        # date was parseable.
        earliest = _pick_earliest_result(asin_results) or any_ok

        display_name = entry.item_name or any_result.product_name or "Not Found"
        price_val = any_ok.price if any_ok and any_ok.price is not None else "Not Found"
        mrp_val = any_ok.mrp if any_ok and any_ok.mrp is not None else "Not Found"
        discount_val = any_ok.discount_percent if any_ok else "N/A"
        bsr_val = any_ok.bsr if any_ok else "N/A"
        availability_val = (any_ok.in_stock if any_ok else (any_result.in_stock or "Not Found"))
        rating_val = any_ok.rating if any_ok else "Not Found"
        reviews_val = any_ok.reviews if any_ok else "Not Found"
        seller_val = any_ok.seller if any_ok else "Amazon"
        earliest_delivery_val = (earliest.delivery_date if earliest else "N/A") or "N/A"
        free_delivery_val = (earliest.free_delivery if earliest else "N/A") or "N/A"
        scraped_val = any_result.scraped_at or ""

        product_url = asin_url(asin)

        row = [
            entry.category or "",
            display_name,
            entry.lapcare_code or "",
            product_url,           # ASIN Link — will be made clickable below
            price_val,
            mrp_val,
            discount_val,
            bsr_val,
            availability_val,
            rating_val,
            reviews_val,
            seller_val,
            earliest_delivery_val,
            free_delivery_val,
            scraped_val,
        ]

        # Append per-pincode availability columns
        for pincode in pincodes:
            pc_result = asin_results.get(pincode)
            if pc_result is None:
                row.append("Pending")
            else:
                row.append(_pincode_cell_value(pc_result))

        ws.append(row)
        row_num = ws.max_row

        # Make ASIN Link a clickable hyperlink
        try:
            link_cell = ws.cell(row=row_num, column=_ASIN_LINK_COL)
            link_cell.hyperlink = product_url
            link_cell.style = "Hyperlink"
            link_cell.value = product_url
        except Exception:
            pass

        _style_pivoted_row(ws, row_num, asin_results, len(FIXED_HEADERS), pincodes)

    # Failed sheet
    ws_failed = wb.create_sheet("Failed")
    ws_failed.append(["ASIN", "Pincode", "City", "Failure Reason", "Timestamp"])
    format_failed_header(ws_failed)
    for fr in failed_rows:
        ws_failed.append(fr)

    # Summary sheet placeholder (caller fills it)
    wb.create_sheet("Summary")

    return wb


def autofit_columns(ws, max_col: int) -> None:
    widths = [0] * (max_col + 1)
    for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=True):
        for i, val in enumerate(row, start=1):
            s = "" if val is None else str(val)
            widths[i] = max(widths[i], min(60, len(s) + 2))
    for i in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, widths[i])


def estimate_delivery_days(delivery_text: str) -> Optional[int]:
    t = (delivery_text or "").strip()
    if not t or t in ("Not Found", "Set pincode manually"):
        return None
    low = t.lower()
    if "today" in low:
        return 0
    if "tomorrow" in low:
        return 1

    m = re.search(r"(\d{1,2})(?:\s*[-–]\s*\d{1,2})?\s+([A-Za-z]{3,})", t)
    if not m:
        return None
    day = int(m.group(1))
    mon = m.group(2)[:3].title()
    months = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
    if mon not in months:
        return None

    today = datetime.now().date()
    year = today.year
    target = datetime(year, months[mon], day).date()
    if target < today - timedelta(days=1):
        target = datetime(year + 1, months[mon], day).date()
    return max(0, (target - today).days)


def write_summary_sheet(
    wb: Workbook,
    totals: Dict[str, object],
    started_at: datetime,
    finished_at: datetime,
) -> None:
    ws = wb["Summary"]
    ws.delete_rows(1, ws.max_row)

    total_asins = int(totals["total_asins"])
    total_combos = int(totals["total_combos"])
    success = int(totals["success"])
    failed = int(totals["failed"])

    avg_price = ""
    if totals["price_count"] > 0:
        avg_price = round(totals["price_sum"] / totals["price_count"], 2)
    else:
        avg_price = "N/A"

    avg_rating = ""
    if totals["rating_count"] > 0:
        avg_rating = round(totals["rating_sum"] / totals["rating_count"], 2)
    else:
        avg_rating = "N/A"

    time_taken = finished_at - started_at
    rate = (success / total_combos * 100) if total_combos else 0

    ws.append(["Metric", "Value"])
    ws.append(["Total ASINs scraped", total_asins])
    ws.append(["Total combinations attempted", total_combos])
    ws.append(["Success rate %", round(rate, 2)])
    ws.append(["Out of stock count", int(totals["out_of_stock"])])
    ws.append(["Average price", avg_price])
    ws.append(["Average rating", avg_rating])
    ws.append(["Scrape date and time", finished_at.strftime("%d %b %Y, %I:%M %p")])
    ws.append(["Time taken", str(time_taken).split(".")[0]])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:B1"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28


# =====================================================
# Email delivery
# =====================================================


def send_email_report(settings: Dict[str, object], xlsx_path: Path, totals: Dict[str, object], started_at: datetime, finished_at: datetime) -> bool:
    date_tag = datetime.now().strftime("%d %b %Y")
    subj = str(settings["EMAIL_SUBJECT"]).replace("{date}", date_tag)

    total_asins = int(totals["total_asins"])
    total_combos = int(totals["total_combos"])
    success = int(totals["success"])
    failed = int(totals["failed"])
    pincodes_checked = int(totals["pincodes_checked"])
    time_taken = finished_at - started_at
    rate = (success / total_combos * 100) if total_combos else 0

    body = (
        "Hi,\n\n"
        "Your Amazon report is ready.\n\n"
        "Summary:\n"
        f"- Date: {date_tag}\n"
        f"- ASINs scraped: {total_asins}\n"
        f"- Pincodes checked: {pincodes_checked}\n"
        f"- Total combinations: {total_combos}\n"
        f"- Successful: {success} ({rate:.1f}%)\n"
        f"- Failed: {failed}\n"
        f"- Time taken: {str(time_taken).split('.')[0]}\n\n"
        "Please find the Excel report attached.\n\n"
        'Failed combinations are listed in the \n'
        '"Failed" sheet of the Excel file.\n'
    )

    msg = EmailMessage()
    msg["From"] = settings["EMAIL_FROM"]
    msg["To"] = settings["EMAIL_TO"]
    msg["Subject"] = subj
    msg.set_content(body)

    try:
        data = xlsx_path.read_bytes()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=xlsx_path.name,
        )
    except Exception:
        print("Could not attach the Excel file to email. Email will be skipped.")
        return False

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(settings["EMAIL_FROM"], settings["EMAIL_PASSWORD"])
            smtp.send_message(msg)
        print(f"Report emailed to {settings['EMAIL_TO']} ✅")
        return True
    except Exception:
        print("Could not send email. Please check email settings in config.py.")
        return False


# =====================================================
# Text / file parsing helpers (used by gui.py)
# =====================================================


def parse_asins_from_text(text: str) -> List[ASINEntry]:
    entries: List[ASINEntry] = []
    current_category = ""
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            current_category = line[1:].strip()
            continue
        line = line.split("#", 1)[0].strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(",")]
        asin = parts[0]
        item_name = parts[1] if len(parts) > 1 else ""
        lapcare_code = parts[2] if len(parts) > 2 else ""
        if is_valid_asin(asin):
            entries.append(ASINEntry(asin=asin, item_name=item_name,
                                     lapcare_code=lapcare_code, category=current_category))
        else:
            print(f"Skipped invalid ASIN: {asin!r}")
    return entries


def read_pincodes_from_file(path: Path) -> Dict[str, str]:
    pincodes: Dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "," in line:
            p, c = [x.strip() for x in line.split(",", 1)]
            if p.isdigit() and len(p) == 6 and c:
                pincodes[p] = c
    return pincodes


def parse_pincodes_from_text(text: str) -> Dict[str, str]:
    pincodes: Dict[str, str] = {}
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "," in line:
            p, c = [x.strip() for x in line.split(",", 1)]
            if p.isdigit() and len(p) == 6 and c:
                pincodes[p] = c
    return pincodes


# =====================================================
# Parallel worker entry point (called by gui.py)
# =====================================================


def run_worker(
    worker_id: int,
    pincodes: Dict[str, str],
    asin_entries: List[ASINEntry],
    settings: Dict[str, object],
    base_dir_str: str,
    msg_queue,
) -> None:
    """Runs in a separate process. Scrapes the given pincodes subset and sends
    progress messages to msg_queue so the GUI can update in real time."""
    base_dir = Path(base_dir_str)
    worker_dir = base_dir / "progress" / f"worker_{worker_id}"
    worker_dir.mkdir(parents=True, exist_ok=True)

    logger = setup_logging(worker_dir)
    results: Dict[str, Dict[str, ScrapeResult]] = {}
    failed_rows: List[List] = []

    def qput(msg: dict) -> None:
        try:
            msg_queue.put_nowait(msg)
        except Exception:
            pass

    total_combos = len(asin_entries) * len(pincodes)
    done = 0
    driver = None

    try:
        driver = build_driver(bool(settings["HEADLESS"]), logger, base_dir, worker_id=worker_id)
        open_homepage(driver, logger)

        for pincode, city in pincodes.items():
            time.sleep(random.uniform(float(settings["MIN_DELAY"]), float(settings["MAX_DELAY"])))
            ok = set_pincode(driver, pincode, city, logger)
            # BUG-FIX: Confirm pincode applied before scraping ASINs — fixes delivery date copy bug
            if ok:  # BUG-FIX
                confirmed = wait_for_pincode_confirmation(driver, pincode)  # BUG-FIX
                if not confirmed:  # BUG-FIX
                    qput({"type": "progress", "worker": worker_id, "done": done,  # BUG-FIX
                          "total": total_combos, "status": "WARN",  # BUG-FIX
                          "msg": f"⚠️  W{worker_id} Pincode {pincode} ({city}) not confirmed — dates may be wrong"})  # BUG-FIX

            if not ok:
                ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                for entry in asin_entries:
                    done += 1
                    res = ScrapeResult(
                        asin=entry.asin, product_name="Not Found",
                        mrp=None, price=None, discount_percent="N/A",
                        pincode=pincode, city=city,
                        in_stock="Not Found", delivery_date="Pincode Failed",
                        free_delivery="N/A", seller="Amazon",
                        rating="Not Found", reviews="Not Found", bsr="N/A",
                        product_url=asin_url(entry.asin), scraped_at=ts,
                        status="PINCODE_FAILED", failure_reason="Pincode Failed",
                    )
                    results.setdefault(entry.asin, {})[pincode] = res
                    failed_rows.append([entry.asin, pincode, city, "Pincode Failed", ts])
                    qput({"type": "progress", "worker": worker_id, "done": done,
                          "total": total_combos, "status": "PINCODE_FAILED",
                          "msg": f"❌ W{worker_id} Pincode {pincode} ({city}) failed"})
                continue

            for entry in asin_entries:
                done += 1
                time.sleep(random.uniform(float(settings["MIN_DELAY"]), float(settings["MAX_DELAY"])))
                try:
                    res = scrape_one(driver, entry.asin, pincode, city, logger)
                    if res.failure_reason == "CAPTCHA":
                        qput({"type": "progress", "worker": worker_id, "done": done,
                              "total": total_combos, "status": "CAPTCHA",
                              "msg": f"⚠️  W{worker_id} CAPTCHA detected — pausing up to 5 min"})
                        # Polls + emits countdown ticks to the GUI via msg_queue
                        # so the worker doesn't look frozen.
                        pause_for_captcha(
                            driver, logger, worker_dir, done, [], msg_queue=msg_queue
                        )
                        # CLAUDE.md: do NOT reuse a stale driver after CAPTCHA pause
                        driver = _recycle_driver(
                            driver, settings, logger, base_dir, worker_id=worker_id
                        )
                        if set_pincode(driver, pincode, city, logger):
                            wait_for_pincode_confirmation(driver, pincode)
                        res = scrape_one(driver, entry.asin, pincode, city, logger)
                    if res.status != "OK":
                        time.sleep(random.uniform(5, 15))
                        try:
                            driver.refresh()
                        except Exception:
                            pass
                        res = scrape_one(driver, entry.asin, pincode, city, logger)

                    # Periodic browser recycle + crash recovery, AFTER the scrape.
                    alive, should_recycle = check_browser_health(driver, logger)
                    if (not alive) or should_recycle:
                        driver = _recycle_driver(
                            driver, settings, logger, base_dir, worker_id=worker_id
                        )
                        if set_pincode(driver, pincode, city, logger):
                            wait_for_pincode_confirmation(driver, pincode)
                except Exception as exc:
                    logger.exception("scrape error in worker")
                    ts2 = datetime.now().strftime("%d %b %Y, %I:%M %p")
                    res = ScrapeResult(
                        asin=entry.asin, product_name="Not Found",
                        mrp=None, price=None, discount_percent="N/A",
                        pincode=pincode, city=city,
                        in_stock="Check on Amazon", delivery_date="Not Found",
                        free_delivery="N/A", seller="Amazon",
                        rating="Not Found", reviews="Not Found", bsr="N/A",
                        product_url=asin_url(entry.asin), scraped_at=ts2,
                        status="FAILED", failure_reason=str(exc)[:120],
                    )

                results.setdefault(entry.asin, {})[pincode] = res
                if res.status != "OK":
                    failed_rows.append([entry.asin, pincode, city,
                                        res.failure_reason or "Failed", res.scraped_at])

                icon = "✅" if res.status == "OK" else "❌"
                price_str = f"₹{int(res.price)}" if res.price else "NA"
                delivery_str = (res.delivery_date or "N/A")[:18]
                qput({
                    "type": "progress",
                    "worker": worker_id,
                    "done": done,
                    "total": total_combos,
                    "status": res.status,
                    "msg": (
                        f"{icon} W{worker_id} [{done:04d}/{total_combos}] "
                        f"{entry.asin} | {city:<10} | {price_str} | "
                        f"{delivery_str} | {res.in_stock}"
                    ),
                })

        # Serialize results into plain dicts for queue transport
        serialized: Dict = {}
        for asin, pc_dict in results.items():
            serialized[asin] = {}
            for pc, r in pc_dict.items():
                serialized[asin][pc] = {
                    "asin": r.asin, "product_name": r.product_name,
                    "mrp": r.mrp, "price": r.price,
                    "discount_percent": r.discount_percent,
                    "pincode": r.pincode, "city": r.city,
                    "in_stock": r.in_stock, "delivery_date": r.delivery_date,
                    "free_delivery": r.free_delivery, "seller": r.seller,
                    "rating": r.rating, "reviews": r.reviews, "bsr": r.bsr,
                    "product_url": r.product_url, "scraped_at": r.scraped_at,
                    "status": r.status, "failure_reason": r.failure_reason,
                }
        qput({"type": "done", "worker": worker_id,
              "results": serialized, "failed_rows": failed_rows})

    except Exception as e:
        logger.exception("Worker fatal error")
        qput({"type": "error", "worker": worker_id, "msg": str(e)})
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# =====================================================
# Main flow
# =====================================================


def estimate_time(total_combos: int) -> str:
    # ~28 seconds per combo: 5s avg delay + 3s page load + 3s settle + 5s pincode + 2s extraction
    seconds = int(total_combos * 28)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    if hours <= 0:
        return f"~{minutes} minutes"
    return f"~{hours} hours {minutes} minutes"


def open_file_cross_platform(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        pass


def main() -> None:
    # Handle --version before any startup side-effects (lock, Chrome, etc.)
    if "--version" in sys.argv:
        print(f"AmazonScraper {__version__} | Python {sys.version.split()[0]}")
        try:
            import undetected_chromedriver as _uc
            print(f"undetected-chromedriver {_uc.__version__}")
        except Exception:
            pass
        try:
            import selenium as _sel
            print(f"selenium {_sel.__version__}")
        except Exception:
            pass
        try:
            import psutil as _ps
            print(f"psutil {_ps.__version__}")
        except Exception:
            pass
        try:
            import openpyxl as _xl
            print(f"openpyxl {_xl.__version__}")
        except Exception:
            pass
        sys.exit(0)

    print_header()

    # BUG-FIX: Capture run timestamp ONCE at startup so Excel filename and log
    # filename share the same tag, and so two runs in the same minute don't collide.
    RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")  # BUG-FIX

    base_dir = Path(__file__).resolve().parent
    folders = ensure_folders(base_dir)

    # Startup sequence per CLAUDE.md:
    # 1. Clean up stale Chrome temp dirs older than 24h
    cleanup_old_chrome_dirs()
    # 2. Delete logs older than 30 days
    cleanup_old_logs(folders["logs"])
    # 3. Acquire PID lock — exits if another instance is already running
    acquire_lock(base_dir)

    logger = setup_logging(folders["logs"], timestamp=RUN_TIMESTAMP)  # BUG-FIX
    settings = validate_config()

    # Register signal handlers so Ctrl+C and SIGTERM save progress and clean up
    signal.signal(signal.SIGINT, _handle_exit_signal)
    signal.signal(signal.SIGTERM, _handle_exit_signal)

    asin_entries = read_asins(base_dir / "asins.txt")
    pincodes = read_pincodes(base_dir)

    # Test mode
    test_mode = "--test" in sys.argv
    if test_mode:
        asin_entries = asin_entries[:3]
        pincodes = dict(list(pincodes.items())[:2])
        print("TEST MODE - Scraping 6 combinations only")

    asins = [e.asin for e in asin_entries]
    total_combos = len(asin_entries) * len(pincodes)
    print(f"Total combinations: {total_combos:,}")
    print(f"Estimated time: {estimate_time(total_combos)}")

    # Resume support
    progress = load_progress(folders["progress"])
    prev_completed: List[List[str]] = progress.get("completed_combinations", []) if isinstance(progress, dict) else []
    prev_done = len(prev_completed)

    completed_set = set(tuple(x) for x in prev_completed if isinstance(x, list) and len(x) == 2)
    resume = False
    if prev_done > 0:
        resume = ask_resume(prev_done, total_combos)
        if not resume:
            completed_set = set()
            prev_completed = []
            save_progress(folders["progress"], 0, [])

    # Load existing results cache when resuming
    results_cache: Dict[str, Dict[str, ScrapeResult]] = {}
    failed_rows: List[List] = []
    if resume:
        results_cache = load_results_cache(folders["progress"])

    xlsx_path = resolve_output_path(settings, timestamp=RUN_TIMESTAMP)  # BUG-FIX

    started_at = datetime.now()

    totals = {
        "total_asins": len(asin_entries),
        "total_combos": total_combos,
        "pincodes_checked": len(pincodes),
        "success": 0,
        "failed": 0,
        "out_of_stock": 0,
        "price_sum": 0.0,
        "price_count": 0,
        "rating_sum": 0.0,
        "rating_count": 0,
    }

    driver: Optional[Chrome] = None
    completed_list: List[List[str]] = list(prev_completed) if resume else []
    done_counter = len(completed_set)
    success_counter = 0
    failed_counter = 0

    # Give signal handler access to live progress state
    _signal_progress_state["progress_dir"] = folders["progress"]
    _signal_progress_state["completed_list"] = completed_list
    _signal_progress_state["done_counter"] = done_counter

    try:
        driver = build_driver(bool(settings["HEADLESS"]), logger, base_dir)
        open_homepage(driver, logger)

        for pincode, city in pincodes.items():
            human_delay(settings)
            ok = set_pincode(driver, pincode, city, logger)
            # BUG-FIX: Added confirmation wait — this is what fixes date copying across pincodes
            if ok:  # BUG-FIX
                confirmed = wait_for_pincode_confirmation(driver, pincode)  # BUG-FIX
                if not confirmed:  # BUG-FIX
                    print(f"⚠️  Pincode {pincode} ({city}) may not have applied — delivery dates may be wrong.")  # BUG-FIX
            if not ok:
                # Mark all ASINs for this pincode as failed
                for entry in asin_entries:
                    combo = (entry.asin, pincode)
                    if combo in completed_set:
                        continue
                    done_counter += 1
                    completed_set.add(combo)
                    completed_list.append([entry.asin, pincode])
                    ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                    res = ScrapeResult(
                        asin=entry.asin,
                        product_name="Not Found",
                        mrp=None,
                        price=None,
                        discount_percent="N/A",
                        pincode=pincode,
                        city=city,
                        in_stock="Not Found",
                        delivery_date="Pincode Failed",
                        free_delivery="N/A",
                        seller="Amazon",
                        rating="Not Found",
                        reviews="Not Found",
                        bsr="N/A",
                        product_url=asin_url(entry.asin),
                        scraped_at=ts,
                        status="PINCODE_FAILED",
                        failure_reason="Pincode Failed",
                    )
                    results_cache.setdefault(entry.asin, {})[pincode] = res
                    failed_rows.append([entry.asin, pincode, city, "Pincode Failed", ts])
                    failed_counter += 1
                    totals["failed"] += 1

                save_progress(folders["progress"], done_counter, completed_list)
                save_results_cache(folders["progress"], results_cache)
                continue

            for entry in asin_entries:
                combo = (entry.asin, pincode)
                if combo in completed_set:
                    continue

                idx_display = done_counter + 1
                res = None
                try:
                    human_delay(settings)
                    res, driver = scrape_with_smart_retry(
                        driver,
                        entry.asin,
                        pincode,
                        city,
                        logger,
                        folders["progress"],
                        idx_display,
                        total_combos,
                        completed_list,
                        settings,
                        base_dir,
                    )

                    # Periodic browser recycle + crash recovery. Done AFTER the
                    # scrape so we never throw away an in-flight page.
                    alive, should_recycle = check_browser_health(driver, logger)
                    if (not alive) or should_recycle:
                        driver = _recycle_driver(driver, settings, logger, base_dir)
                        if set_pincode(driver, pincode, city, logger):
                            wait_for_pincode_confirmation(driver, pincode)

                    if res.status == "OK":
                        success_counter += 1
                        totals["success"] += 1
                        if res.price is not None:
                            totals["price_sum"] += float(res.price)
                            totals["price_count"] += 1
                        if res.rating not in ("Not Found", ""):
                            try:
                                totals["rating_sum"] += float(res.rating)
                                totals["rating_count"] += 1
                            except Exception:
                                pass
                        if (res.in_stock or "").lower().startswith("out of stock"):
                            totals["out_of_stock"] += 1
                    else:
                        failed_counter += 1
                        totals["failed"] += 1
                        failed_rows.append([entry.asin, pincode, city, res.failure_reason or "Failed", res.scraped_at])

                except Exception:
                    logger.exception("Unexpected scrape error")
                    failed_counter += 1
                    totals["failed"] += 1
                    ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                    res = ScrapeResult(
                        asin=entry.asin,
                        product_name="Not Found",
                        mrp=None,
                        price=None,
                        discount_percent="N/A",
                        pincode=pincode,
                        city=city,
                        in_stock="Check on Amazon",
                        delivery_date="Not Found",
                        free_delivery="N/A",
                        seller="Amazon",
                        rating="Not Found",
                        reviews="Not Found",
                        bsr="N/A",
                        product_url=asin_url(entry.asin),
                        scraped_at=ts,
                        status="FAILED",
                        failure_reason="Unexpected error",
                    )
                    failed_rows.append([entry.asin, pincode, city, "Unexpected error", ts])
                    print(f"❌ [{idx_display:04d}/{total_combos}] {entry.asin} | {city:<10} | Failed")

                # Store in results cache (keyed by asin → pincode)
                if res is not None:
                    results_cache.setdefault(entry.asin, {})[pincode] = res

                done_counter += 1
                _signal_progress_state["done_counter"] = done_counter
                completed_set.add(combo)
                completed_list.append([entry.asin, pincode])

                # Save progress every 10 scrapes
                if done_counter % 10 == 0:
                    save_progress(folders["progress"], done_counter, completed_list)
                    save_results_cache(folders["progress"], results_cache)

                # Rebuild Excel every 50 scrapes
                if done_counter % 50 == 0:
                    try:
                        wb_interim = build_pivoted_excel(results_cache, asin_entries, pincodes, xlsx_path, failed_rows, logger)
                        wb_interim.save(xlsx_path)
                    except Exception:
                        logger.exception("Interim Excel save failed (non-fatal)")

                # Summary line every 10 scrapes
                if done_counter % 10 == 0:
                    elapsed = datetime.now() - started_at
                    pct = (done_counter / total_combos * 100) if total_combos else 0.0
                    per = elapsed.total_seconds() / max(1, done_counter)
                    rem = int(per * max(0, total_combos - done_counter))
                    rem_h = rem // 3600
                    rem_m = (rem % 3600) // 60
                    print(
                        f"Progress: {done_counter}/{total_combos} ({pct:.1f}%) | "
                        f"Success: {success_counter} | Failed: {failed_counter} | "
                        f"Time elapsed: {int(elapsed.total_seconds()//60)} min | "
                        f"Estimated remaining: {rem_h}h {rem_m}m"
                    )

        # Final Excel build with all results
        finished_at = datetime.now()
        try:
            wb = build_pivoted_excel(results_cache, asin_entries, pincodes, xlsx_path, failed_rows, logger)
            autofit_columns(wb["Results"], len(FIXED_HEADERS) + len(pincodes))
            autofit_columns(wb["Failed"], 5)
            write_summary_sheet(wb, totals, started_at, finished_at)
            wb.save(xlsx_path)
        except Exception:
            logger.exception("Final Excel build failed")

        # Email (optional)
        email_sent = False
        if settings["SEND_EMAIL"]:
            email_sent = send_email_report(settings, xlsx_path, totals, started_at, finished_at)

        total = total_combos
        succ = int(totals["success"])
        fail = int(totals["failed"])
        rate = (succ / total * 100) if total else 0
        time_taken = finished_at - started_at

        print("\n" + "=" * 32)
        print("SCRAPING COMPLETE")
        print("=" * 32)
        print(f"Total combinations: {total:,}")
        print(f"Successful:         {succ:,} ({rate:.1f}%)")
        print(f"Failed:             {fail:,} ({(fail/total*100 if total else 0):.1f}%)")
        print(f"Time taken:         {str(time_taken).split('.')[0]}")
        print(f"Report saved to:    {xlsx_path}")
        if settings["SEND_EMAIL"]:
            print(f"Email sent to:      {settings['EMAIL_TO'] if email_sent else 'Not sent'}")
        print("=" * 32)

        open_file_cross_platform(xlsx_path)

        # Clean up progress files for a fresh start next time
        try:
            pf = progress_file(folders["progress"])
            if pf.exists():
                pf.unlink()
            rc = folders["progress"] / "results_cache.json"
            if rc.exists():
                rc.unlink()
        except Exception:
            pass

    except KeyboardInterrupt:
        try:
            save_progress(folders["progress"], done_counter, completed_list)
            save_results_cache(folders["progress"], results_cache)
            try:
                wb_partial = build_pivoted_excel(results_cache, asin_entries, pincodes, xlsx_path, failed_rows, logger)
                wb_partial.save(xlsx_path)
            except Exception:
                logger.exception("Excel save failed during safe stop")
        except Exception:
            pass
        print("\nStopping safely. Your progress has been saved. You can run again to resume.")
    except SystemExit:
        raise
    except Exception:
        logger.exception("Fatal error")
        try:
            save_progress(folders["progress"], done_counter, completed_list)
            save_results_cache(folders["progress"], results_cache)
            try:
                wb_partial = build_pivoted_excel(results_cache, asin_entries, pincodes, xlsx_path, failed_rows, logger)
                wb_partial.save(xlsx_path)
            except Exception:
                logger.exception("Excel save failed after fatal error")
        except Exception:
            pass
        print("Something went wrong, but your progress was saved. Please try again.")
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
